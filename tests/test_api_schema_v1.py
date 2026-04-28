from __future__ import annotations
"""Contract tests for frozen API schema v1."""

import asyncio
import io
import json
import os
from pathlib import Path
from typing import Any
from uuid import uuid4

import pytest
from openpyxl import Workbook, load_workbook
from pydantic import ValidationError

from bills_analysis.integrations.local_backend import LocalPipelineBackend
from bills_analysis.models.api_requests import CreateBatchRequest, CreateBatchUploadForm, MergeRequest
from bills_analysis.models.internal import BatchRecord


async def _append_file_done_event(events: list[dict[str, Any]], event: dict[str, Any]) -> None:
    """Collect backend per-file completion callbacks in tests."""

    events.append(event)


def _get_test_client_and_app():
    """Lazily import FastAPI app to allow model-only tests without web deps."""

    pytest.importorskip("fastapi")
    from fastapi.testclient import TestClient

    os.environ["RUN_INLINE_WORKER"] = "false"
    from bills_analysis.api.main import app

    return TestClient, app


def test_create_batch_request_valid() -> None:
    """CreateBatchRequest accepts valid daily payload."""

    req = CreateBatchRequest(
        type="daily",
        run_date="04/02/2026",
        inputs=[{"path": "a.pdf", "category": "bar"}],
        metadata={},
    )
    assert req.type == "daily"
    assert req.run_date == "04/02/2026"


def test_statistics_summary_defaults() -> None:
    """StatisticsSummary should default all amounts to zero."""

    from bills_analysis.models.api_responses import StatisticsSummary

    summary = StatisticsSummary()
    assert summary.revenue_brutto == 0
    assert summary.daily_expense_brutto == 0
    assert summary.office_expense_brutto == 0
    assert summary.profit_brutto == 0


def test_statistics_summary_rejects_extra_field() -> None:
    """StatisticsSummary must reject unknown fields via StrictModel."""

    from bills_analysis.models.api_responses import StatisticsSummary

    with pytest.raises(ValidationError):
        StatisticsSummary(revenue_brutto=1.0, unknown_field="x")


def test_daily_statistics_point_requires_date() -> None:
    """DailyStatisticsPoint requires a date string."""

    from bills_analysis.models.api_responses import DailyStatisticsPoint

    point = DailyStatisticsPoint(date="2025-11-01")
    assert point.date == "2025-11-01"
    assert point.revenue_brutto == 0


def test_office_type_breakdown_share() -> None:
    """OfficeTypeBreakdown stores computed share as float."""

    from bills_analysis.models.api_responses import OfficeTypeBreakdown

    breakdown = OfficeTypeBreakdown(type="Miete", brutto=5000.0, count=2, share=0.45)
    assert breakdown.share == 0.45


def test_office_statistics_row_optional_fields() -> None:
    """OfficeStatisticsRow allows null date and name."""

    from bills_analysis.models.api_responses import OfficeStatisticsRow

    row = OfficeStatisticsRow(type="Miete")
    assert row.date is None
    assert row.name is None
    assert row.brutto == 0


def test_monthly_statistics_response_schema_version() -> None:
    """MonthlyStatisticsResponse must carry schema_version == v1."""

    from bills_analysis.models.api_responses import MonthlyStatisticsResponse, StatisticsSummary

    response = MonthlyStatisticsResponse(summary=StatisticsSummary(), warnings=[])
    assert response.schema_version == "v1"
    assert response.daily_series == []
    assert response.office_by_type == []
    assert response.office_rows == []
    assert response.warnings == []


def test_create_batch_request_alias_batch_type_is_accepted() -> None:
    """Backward-compatible `batch_type` alias should still parse."""

    req = CreateBatchRequest(
        batch_type="office",
        run_date="04/02/2026",
        inputs=[{"path": "b.pdf", "category": "office"}],
        metadata={},
    )
    assert req.type == "office"


def test_create_batch_request_invalid_run_date_rejected() -> None:
    """Invalid run_date format must be rejected by regex validation."""

    with pytest.raises(ValidationError):
        CreateBatchRequest(
            type="daily",
            run_date="2026-02-04",
            inputs=[{"path": "a.pdf", "category": "bar"}],
            metadata={},
        )


def test_create_batch_request_empty_inputs_rejected() -> None:
    """Empty inputs list must fail validation."""

    with pytest.raises(ValidationError):
        CreateBatchRequest(type="daily", run_date="04/02/2026", inputs=[], metadata={})


def test_create_batch_request_extra_field_rejected() -> None:
    """Unknown request field must fail because extra=forbid."""

    with pytest.raises(ValidationError):
        CreateBatchRequest(
            type="daily",
            run_date="04/02/2026",
            inputs=[{"path": "a.pdf", "category": "bar"}],
            metadata={},
            foo="bar",
        )


def test_create_batch_request_invalid_type_rejected() -> None:
    """Unknown batch type enum value must be rejected."""

    with pytest.raises(ValidationError):
        CreateBatchRequest(
            type="unknown",
            run_date="04/02/2026",
            inputs=[{"path": "a.pdf", "category": "bar"}],
            metadata={},
        )


def test_merge_request_invalid_mode_rejected() -> None:
    """Unknown merge mode must fail validation."""

    with pytest.raises(ValidationError):
        MergeRequest(mode="upsert")


def test_create_batch_upload_form_invalid_run_date_rejected() -> None:
    """CreateBatchUploadForm must reject non-DD/MM/YYYY run_date."""

    with pytest.raises(ValidationError):
        CreateBatchUploadForm(type="daily", run_date="2026-02-04", metadata={})


def test_api_contract_v1_endpoints() -> None:
    """End-to-end API contract checks for v1 routes and response shapes."""

    TestClient, app = _get_test_client_and_app()
    with TestClient(app) as client:
        create_res = client.post(
            "/v1/batches",
            headers={"Origin": "http://127.0.0.1:5173"},
            json={
                "type": "daily",
                "run_date": "04/02/2026",
                "inputs": [{"path": "a.pdf", "category": "bar"}],
                "metadata": {},
            },
        )
        assert create_res.status_code == 200
        assert create_res.headers.get("access-control-allow-origin") == "http://127.0.0.1:5173"
        created = create_res.json()
        assert created["schema_version"] == "v1"
        assert created["inputs"][0]["status"] == "queued"
        assert created["inputs"][0]["error"] is None
        batch_id = created["batch_id"]

        get_res = client.get(f"/v1/batches/{batch_id}")
        assert get_res.status_code == 200
        got = get_res.json()
        assert got["schema_version"] == "v1"
        assert got["batch_id"] == batch_id

        review_res = client.put(
            f"/v1/batches/{batch_id}/review",
            json={"rows": [{"category": "bar", "filename": "a.pdf", "result": {"brutto": "1.0"}}]},
        )
        assert review_res.status_code == 200
        reviewed = review_res.json()
        assert reviewed["review_rows_count"] == 1

        merge_res = client.post(
            f"/v1/batches/{batch_id}/merge",
            json={"mode": "overwrite", "monthly_excel_path": "outputs/monthly.xlsx", "metadata": {}},
        )
        assert merge_res.status_code == 200
        merged_task = merge_res.json()
        assert merged_task["schema_version"] == "v1"
        assert merged_task["batch_id"] == batch_id
        assert merged_task["task_type"] == "merge_batch"

        list_res = client.get("/v1/batches")
        assert list_res.status_code == 200
        list_body = list_res.json()
        assert list_body["schema_version"] == "v1"
        assert isinstance(list_body["items"], list)
        assert list_body["total"] >= 1


def test_office_receiver_options_endpoint_contract() -> None:
    """Office receiver options endpoint should return v1 envelope with city options."""

    TestClient, app = _get_test_client_and_app()
    with TestClient(app) as client:
        res = client.get("/v1/batches/office-receiver-options")
        assert res.status_code == 200
        body = res.json()
        assert body["schema_version"] == "v1"
        assert isinstance(body["default_city"], str)
        assert isinstance(body["options"], list)
        assert len(body["options"]) >= 1
        first = body["options"][0]
        assert isinstance(first["city"], str)
        assert isinstance(first["receiver_name"], str)
        assert isinstance(first["receiver_address"], str)


def test_cors_preflight_options_for_create_batch() -> None:
    """CORS preflight OPTIONS for create-batch endpoint should not return 405."""

    TestClient, app = _get_test_client_and_app()
    with TestClient(app) as client:
        res = client.options(
            "/v1/batches",
            headers={
                "Origin": "http://127.0.0.1:5173",
                "Access-Control-Request-Method": "POST",
            },
        )
        assert res.status_code == 200
        assert res.headers.get("access-control-allow-origin") == "http://127.0.0.1:5173"
        assert "POST" in (res.headers.get("access-control-allow-methods") or "")


def test_cors_uses_env_allow_origins_for_parallel_frontend_ports(monkeypatch: pytest.MonkeyPatch) -> None:
    """Configured CORS origins should allow a second local frontend port for parallel runs."""

    monkeypatch.setenv(
        "CORS_ALLOW_ORIGINS",
        "http://127.0.0.1:5173,http://127.0.0.1:5174",
    )
    from bills_analysis.api.main import _load_cors_allow_origins

    origins = _load_cors_allow_origins()
    assert origins == ["http://127.0.0.1:5173", "http://127.0.0.1:5174"]


def _make_excel_bytes() -> bytes:
    """Build in-memory workbook bytes for multipart Excel upload tests."""

    wb = Workbook()
    ws = wb.active
    ws.title = "Monthly"
    ws.append(["Datum", "Umsatz Brutto"])
    ws.append(["04/02/2026", 12.34])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_review_rows_and_preview_routes() -> None:
    """Review rows endpoint should return preview_url and preview route should serve PDF."""

    TestClient, app = _get_test_client_and_app()
    with TestClient(app) as client:
        create_res = client.post(
            "/v1/batches",
            json={
                "type": "daily",
                "run_date": "04/02/2026",
                "inputs": [{"path": "a.pdf", "category": "bar"}],
                "metadata": {},
            },
        )
        assert create_res.status_code == 200
        batch_id = create_res.json()["batch_id"]

        preview_path = Path("outputs") / "webapp" / batch_id / "archive" / "bar" / "preview.pdf"
        preview_path.parent.mkdir(parents=True, exist_ok=True)
        preview_path.write_bytes(b"%PDF-1.4\npreview\n%%EOF")

        review_res = client.put(
            f"/v1/batches/{batch_id}/review",
            json={
                "rows": [
                    {
                        "row_id": "row-0001",
                        "filename": "a.pdf",
                        "category": "bar",
                        "result": {"brutto": "1.0"},
                        "score": {"brutto": 0.9},
                        "preview_path": str(preview_path.resolve()),
                    }
                ]
            },
        )
        assert review_res.status_code == 200

        rows_res = client.get(f"/v1/batches/{batch_id}/review-rows")
        assert rows_res.status_code == 200
        body = rows_res.json()
        assert body["batch_id"] == batch_id
        assert len(body["rows"]) == 1
        assert body["rows"][0]["preview_url"]

        preview_res = client.get(body["rows"][0]["preview_url"])
        assert preview_res.status_code == 200
        assert preview_res.headers["content-type"].startswith("application/pdf")


def test_review_submit_keeps_preview_path_when_payload_omits_it() -> None:
    """Review submit should preserve existing preview_path so preview URL stays usable."""

    TestClient, app = _get_test_client_and_app()
    with TestClient(app) as client:
        create_res = client.post(
            "/v1/batches",
            json={
                "type": "daily",
                "run_date": "04/02/2026",
                "inputs": [{"path": "a.pdf", "category": "bar"}],
                "metadata": {},
            },
        )
        assert create_res.status_code == 200
        batch_id = create_res.json()["batch_id"]

        preview_path = Path("outputs") / "webapp" / batch_id / "archive" / "bar" / "preview_keep.pdf"
        preview_path.parent.mkdir(parents=True, exist_ok=True)
        preview_path.write_bytes(b"%PDF-1.4\npreview\n%%EOF")

        seed_review_res = client.put(
            f"/v1/batches/{batch_id}/review",
            json={
                "rows": [
                    {
                        "row_id": "row-0001",
                        "filename": "a.pdf",
                        "category": "bar",
                        "result": {"brutto": "1.0"},
                        "score": {"brutto": 0.9},
                        "preview_path": str(preview_path.resolve()),
                    }
                ]
            },
        )
        assert seed_review_res.status_code == 200

        review_res = client.put(
            f"/v1/batches/{batch_id}/review",
            json={
                "rows": [
                    {
                        "row_id": "row-0001",
                        "filename": "a.pdf",
                        "category": "bar",
                        "result": {"brutto": "2.0"},
                        "score": {"brutto": 0.95},
                    }
                ]
            },
        )
        assert review_res.status_code == 200

        rows_res = client.get(f"/v1/batches/{batch_id}/review-rows")
        assert rows_res.status_code == 200
        body = rows_res.json()
        assert len(body["rows"]) == 1
        assert body["rows"][0]["preview_url"]

        preview_res = client.get(body["rows"][0]["preview_url"])
        assert preview_res.status_code == 200
        assert preview_res.headers["content-type"].startswith("application/pdf")


def test_submit_review_rejects_missing_result_shape() -> None:
    """Review submit should return 422 when row does not provide canonical nested result."""

    TestClient, app = _get_test_client_and_app()
    with TestClient(app) as client:
        create_res = client.post(
            "/v1/batches",
            json={
                "type": "daily",
                "run_date": "04/02/2026",
                "inputs": [{"path": "a.pdf", "category": "bar"}],
                "metadata": {},
            },
        )
        assert create_res.status_code == 200
        batch_id = create_res.json()["batch_id"]
        review_res = client.put(
            f"/v1/batches/{batch_id}/review",
            json={
                "rows": [
                    {
                        "row_id": "row-0001",
                        "filename": "a.pdf",
                        "category": "bar",
                    }
                ]
            },
        )
        assert review_res.status_code == 422
        assert "canonical shape" in review_res.json()["detail"]


def test_submit_review_flattened_fields_rejected() -> None:
    """Flattened top-level review fields must be rejected after compatibility removal."""

    TestClient, app = _get_test_client_and_app()
    with TestClient(app) as client:
        create_res = client.post(
            "/v1/batches",
            json={
                "type": "daily",
                "run_date": "04/02/2026",
                "inputs": [{"path": "a.pdf", "category": "bar"}],
                "metadata": {},
            },
        )
        assert create_res.status_code == 200
        batch_id = create_res.json()["batch_id"]

        review_res = client.put(
            f"/v1/batches/{batch_id}/review",
            json={
                "rows": [
                    {
                        "row_id": "row-0001",
                        "filename": "a.pdf",
                        "category": "bar",
                        "brutto": "12.30",
                        "netto": "10.00",
                        "store_name": "Demo",
                    }
                ]
            },
        )
        assert review_res.status_code == 422
        assert "result must be an object" in review_res.json()["detail"]


def test_submit_review_canonical_shape_persisted_to_review_artifacts() -> None:
    """Canonical nested review payload should persist both review artifact files."""

    TestClient, app = _get_test_client_and_app()
    with TestClient(app) as client:
        create_res = client.post(
            "/v1/batches",
            json={
                "type": "daily",
                "run_date": "04/02/2026",
                "inputs": [{"path": "a.pdf", "category": "bar"}],
                "metadata": {},
            },
        )
        assert create_res.status_code == 200
        batch_id = create_res.json()["batch_id"]

        review_res = client.put(
            f"/v1/batches/{batch_id}/review",
            json={
                "rows": [
                    {
                        "row_id": "row-0001",
                        "filename": "a.pdf",
                        "category": "bar",
                        "result": {
                            "brutto": "12.30",
                            "netto": "10.00",
                            "store_name": "Demo",
                        },
                        "score": {"brutto": 0.91},
                    }
                ]
            },
        )
        assert review_res.status_code == 200

        rows_res = client.get(f"/v1/batches/{batch_id}/review-rows")
        assert rows_res.status_code == 200
        row = rows_res.json()["rows"][0]
        assert row["result"]["brutto"] == "12.30"
        assert row["result"]["netto"] == "10.00"
        assert row["result"]["store_name"] == "Demo"

        review_file = Path("outputs") / "webapp" / batch_id / "review_rows.json"
        assert review_file.exists()
        saved_rows = json.loads(review_file.read_text(encoding="utf-8"))
        assert saved_rows[0]["result"]["brutto"] == "12.30"

        submitted_file = Path("outputs") / "webapp" / batch_id / "review_rows_submitted.json"
        assert submitted_file.exists()
        submitted_rows = json.loads(submitted_file.read_text(encoding="utf-8"))
        assert submitted_rows[0]["result"]["brutto"] == "12.30"


def test_report_error_endpoint_copies_batch_artifacts_and_returns_type_corrections() -> None:
    """Report-error endpoint should snapshot artifacts and return office type correction diffs."""

    TestClient, app = _get_test_client_and_app()
    with TestClient(app) as client:
        create_res = client.post(
            "/v1/batches",
            json={
                "type": "office",
                "run_date": "04/02/2026",
                "inputs": [{"path": "a.pdf", "category": "office"}],
                "metadata": {},
            },
        )
        assert create_res.status_code == 200
        batch_id = create_res.json()["batch_id"]

        batch_out = Path("outputs") / "webapp" / batch_id
        batch_out.mkdir(parents=True, exist_ok=True)
        results_path = batch_out / "results.json"
        results_path.write_text(
            json.dumps(
                {
                    "batch_id": batch_id,
                    "batch_type": "office",
                    "items": [
                        {
                            "row_id": "row-0001",
                            "filename": "a.pdf",
                            "category": "office",
                            "result": {"type": "Service&Andere"},
                        }
                    ],
                },
                ensure_ascii=False,
                indent=2,
            ),
            encoding="utf-8",
        )
        di_dir = batch_out / "di_fields"
        di_dir.mkdir(parents=True, exist_ok=True)
        (di_dir / "row-0001.json").write_text(
            json.dumps({"invoice_no": "A-100"}, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )

        review_res = client.put(
            f"/v1/batches/{batch_id}/review",
            json={
                "rows": [
                    {
                        "row_id": "row-0001",
                        "filename": "a.pdf",
                        "category": "office",
                        "result": {"type": "Miete"},
                        "score": {},
                    }
                ]
            },
        )
        assert review_res.status_code == 200

        report_res = client.post(f"/v1/batches/{batch_id}/report-error")
        assert report_res.status_code == 200
        body = report_res.json()
        assert body["schema_version"] == "v1"
        assert body["status"] == "reported"
        assert len(body["corrections"]) == 1
        assert body["corrections"][0]["row_id"] == "row-0001"
        assert body["corrections"][0]["original_type"] == "Service&Andere"
        assert body["corrections"][0]["corrected_type"] == "Miete"

        dataset_dir = Path("dataset") / "type_errors" / batch_id
        assert (dataset_dir / "results.json").exists()
        assert (dataset_dir / "review_rows_submitted.json").exists()
        assert (dataset_dir / "di_fields" / "row-0001.json").exists()
        summary_path = dataset_dir / "correction_summary.json"
        assert summary_path.exists()
        summary = json.loads(summary_path.read_text(encoding="utf-8"))
        assert summary["corrections"][0]["corrected_type"] == "Miete"


def test_review_rows_not_found_returns_404() -> None:
    """Review rows route should return 404 for unknown batch id."""

    TestClient, app = _get_test_client_and_app()
    with TestClient(app) as client:
        res = client.get("/v1/batches/not-exist/review-rows")
        assert res.status_code == 404


def test_merge_source_local_upload_and_merge_fallback() -> None:
    """Uploading local monthly source should allow merge without request monthly path."""

    TestClient, app = _get_test_client_and_app()
    with TestClient(app) as client:
        create_res = client.post(
            "/v1/batches",
            json={
                "type": "daily",
                "run_date": "04/02/2026",
                "inputs": [{"path": "a.pdf", "category": "bar"}],
                "metadata": {},
            },
        )
        assert create_res.status_code == 200
        batch_id = create_res.json()["batch_id"]

        upload_res = client.post(
            f"/v1/batches/{batch_id}/merge-source/local",
            files={"file": ("monthly.xlsx", _make_excel_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert upload_res.status_code == 200
        upload_body = upload_res.json()
        assert upload_body["source_type"] == "local_excel"
        assert Path(upload_body["monthly_excel_path"]).exists()

        merge_res = client.post(
            f"/v1/batches/{batch_id}/merge",
            json={"mode": "overwrite", "metadata": {}},
        )
        assert merge_res.status_code == 200
        assert merge_res.json()["task_type"] == "merge_batch"


def test_merge_without_monthly_path_is_accepted_for_auto_template() -> None:
    """Merge queue endpoint should accept missing monthly path for backend auto-template flow."""

    TestClient, app = _get_test_client_and_app()
    with TestClient(app) as client:
        create_res = client.post(
            "/v1/batches",
            json={
                "type": "daily",
                "run_date": "04/02/2026",
                "inputs": [{"path": "a.pdf", "category": "bar"}],
                "metadata": {},
            },
        )
        assert create_res.status_code == 200
        batch_id = create_res.json()["batch_id"]

        merge_res = client.post(
            f"/v1/batches/{batch_id}/merge",
            json={"mode": "overwrite", "metadata": {}},
        )
        assert merge_res.status_code == 200
        assert merge_res.json()["task_type"] == "merge_batch"


def test_merge_source_local_invalid_file_rejected() -> None:
    """Merge source upload should reject non-Excel file types."""

    TestClient, app = _get_test_client_and_app()
    with TestClient(app) as client:
        create_res = client.post(
            "/v1/batches",
            json={
                "type": "daily",
                "run_date": "04/02/2026",
                "inputs": [{"path": "a.pdf", "category": "bar"}],
                "metadata": {},
            },
        )
        assert create_res.status_code == 200
        batch_id = create_res.json()["batch_id"]
        upload_res = client.post(
            f"/v1/batches/{batch_id}/merge-source/local",
            files={"file": ("monthly.txt", b"hello", "text/plain")},
        )
        assert upload_res.status_code == 400


def test_multipart_upload_daily_with_required_single_zbon() -> None:
    """Daily upload succeeds with exactly one zbon_file and no bar_files."""

    TestClient, app = _get_test_client_and_app()
    with TestClient(app) as client:
        upload_res = client.post(
            "/v1/batches/upload",
            data={"type": "daily", "run_date": "04/02/2026"},
            files={
                "zbon_file": ("zbon.pdf", b"%PDF-1.4\nzbon\n%%EOF", "application/pdf"),
            },
        )
        assert upload_res.status_code == 200
        body = upload_res.json()
        assert body["schema_version"] == "v1"
        assert body["type"] == "daily"
        assert body["status"] == "queued"
        assert body["task_id"]
        assert body["batch_id"]


def test_multipart_upload_daily_with_optional_bar_files() -> None:
    """Daily upload accepts optional bar_files together with required zbon_file."""

    TestClient, app = _get_test_client_and_app()
    with TestClient(app) as client:
        upload_res = client.post(
            "/v1/batches/upload",
            data={"type": "daily", "run_date": "04/02/2026"},
            files=[
                ("zbon_file", ("zbon.pdf", b"%PDF-1.4\nzbon\n%%EOF", "application/pdf")),
                ("bar_files", ("bar1.pdf", b"%PDF-1.4\nbar1\n%%EOF", "application/pdf")),
                ("bar_files", ("bar2.pdf", b"%PDF-1.4\nbar2\n%%EOF", "application/pdf")),
            ],
        )
        assert upload_res.status_code == 200
        body = upload_res.json()
        assert body["type"] == "daily"


def test_multipart_upload_daily_missing_zbon_rejected() -> None:
    """Daily upload must reject requests without zbon_file."""

    TestClient, app = _get_test_client_and_app()
    with TestClient(app) as client:
        upload_res = client.post(
            "/v1/batches/upload",
            data={"type": "daily"},
            files=[("bar_files", ("bar1.pdf", b"%PDF-1.4\nbar1\n%%EOF", "application/pdf"))],
        )
        assert upload_res.status_code == 400
        assert "zbon_file" in upload_res.json()["detail"]


def test_multipart_upload_daily_duplicate_zbon_rejected() -> None:
    """Daily upload must reject duplicate zbon_file parts."""

    TestClient, app = _get_test_client_and_app()
    with TestClient(app) as client:
        upload_res = client.post(
            "/v1/batches/upload",
            data={"type": "daily"},
            files=[
                ("zbon_file", ("zbon_1.pdf", b"%PDF-1.4\nz1\n%%EOF", "application/pdf")),
                ("zbon_file", ("zbon_2.pdf", b"%PDF-1.4\nz2\n%%EOF", "application/pdf")),
            ],
        )
        assert upload_res.status_code == 400
        assert "exactly one zbon_file" in upload_res.json()["detail"]


def test_multipart_upload_office_multi_files_success() -> None:
    """Office upload accepts multiple office_files parts."""

    TestClient, app = _get_test_client_and_app()
    with TestClient(app) as client:
        upload_res = client.post(
            "/v1/batches/upload",
            data={"type": "office", "run_date": "04/02/2026"},
            files=[
                ("office_files", ("office_1.pdf", b"%PDF-1.4\no1\n%%EOF", "application/pdf")),
                ("office_files", ("office_2.pdf", b"%PDF-1.4\no2\n%%EOF", "application/pdf")),
            ],
        )
        assert upload_res.status_code == 200
        body = upload_res.json()
        assert body["type"] == "office"
        assert body["status"] == "queued"


def test_multipart_upload_office_missing_files_rejected() -> None:
    """Office upload must reject requests without office_files."""

    TestClient, app = _get_test_client_and_app()
    with TestClient(app) as client:
        upload_res = client.post(
            "/v1/batches/upload",
            data={"type": "office"},
        )
        assert upload_res.status_code == 400
        assert "office_files" in upload_res.json()["detail"]


def test_multipart_upload_non_pdf_rejected() -> None:
    """Upload endpoint must reject non-PDF files."""

    TestClient, app = _get_test_client_and_app()
    with TestClient(app) as client:
        upload_res = client.post(
            "/v1/batches/upload",
            data={"type": "daily"},
            files={"zbon_file": ("zbon.txt", b"hello", "text/plain")},
        )
        assert upload_res.status_code == 400
        assert "PDF" in upload_res.json()["detail"]


def test_multipart_upload_invalid_metadata_json_rejected() -> None:
    """Upload endpoint must reject invalid metadata_json string."""

    TestClient, app = _get_test_client_and_app()
    with TestClient(app) as client:
        upload_res = client.post(
            "/v1/batches/upload",
            data={"type": "daily", "metadata_json": "{not-json}"},
            files={"zbon_file": ("zbon.pdf", b"%PDF-1.4\nzbon\n%%EOF", "application/pdf")},
        )
        assert upload_res.status_code == 400
        assert "metadata_json" in upload_res.json()["detail"]


def test_multipart_upload_invalid_run_date_rejected() -> None:
    """Upload endpoint must reject invalid run_date format."""

    TestClient, app = _get_test_client_and_app()
    with TestClient(app) as client:
        upload_res = client.post(
            "/v1/batches/upload",
            data={"type": "daily", "run_date": "2026-02-04"},
            files={"zbon_file": ("zbon.pdf", b"%PDF-1.4\nzbon\n%%EOF", "application/pdf")},
        )
        assert upload_res.status_code == 422


def test_local_backend_calls_preprocess_and_extract(monkeypatch: pytest.MonkeyPatch) -> None:
    """Local backend process must invoke preprocess and extraction adapters."""

    called: dict[str, Any] = {"compress": 0, "analyze": 0}

    def fake_compress(pdf_path: Path, *, dest_dir: Path, dpi: int, name_suffix: str) -> Path:
        """Stub compression helper returning an archive path."""

        called["compress"] += 1
        dest_dir.mkdir(parents=True, exist_ok=True)
        archived = dest_dir / f"{pdf_path.stem}_{name_suffix}.pdf"
        archived.write_bytes(b"%PDF-1.4\narchived\n%%EOF")
        return archived

    def fake_analyze(pdf_path: Path, *, model_id: str, return_fields: bool) -> Any:
        """Stub extraction helper returning deterministic payload."""

        called["analyze"] += 1
        payload = {
            "store_name": "Demo Shop",
            "brutto": 12.34,
            "netto": 10.0,
            "total_tax": 2.34,
            "confidence_store_name": 0.9,
            "confidence_brutto": 0.95,
            "confidence_netto": 0.9,
            "confidence_total_tax": 0.8,
        }
        if return_fields:
            return payload, {}
        return payload

    monkeypatch.setattr("bills_analysis.integrations.local_backend._compress_pdf_for_archive", fake_compress)
    monkeypatch.setattr("bills_analysis.integrations.local_backend._analyze_pdf_with_azure", fake_analyze)

    test_root = Path("outputs") / "pytest_tmp" / str(uuid4())
    test_root.mkdir(parents=True, exist_ok=True)
    src_pdf = test_root / "input.pdf"
    src_pdf.write_bytes(b"%PDF-1.4\nsource\n%%EOF\n" + (b"0" * (2 * 1024 * 1024)))
    req = CreateBatchRequest(
        type="daily",
        run_date="04/02/2026",
        inputs=[{"path": str(src_pdf), "category": "zbon"}],
        metadata={},
    )
    batch = BatchRecord.new(req)
    backend = LocalPipelineBackend(root=test_root / "out")

    artifacts = asyncio.run(backend.process_batch(batch))

    assert called["compress"] == 1
    assert called["analyze"] == 1
    assert Path(artifacts["artifacts"]["result_json_path"]).exists()
    assert Path(artifacts["artifacts"]["review_json_path"]).exists()
    organized_root = Path(artifacts["artifacts"]["organized_root"])
    assert organized_root.exists()
    assert organized_root.resolve() == (backend.root.parent / "organized").resolve()
    assert len(artifacts["review_rows"]) == 1

    results_payload = json.loads(Path(artifacts["artifacts"]["result_json_path"]).read_text(encoding="utf-8"))
    first_item = results_payload["items"][0]
    assert first_item.get("organized_path")
    organized_path = Path(first_item["organized_path"])
    assert organized_path.exists()
    assert organized_root.resolve() in organized_path.resolve().parents


def test_local_backend_persists_office_di_fields_artifact(monkeypatch: pytest.MonkeyPatch) -> None:
    """Office processing should persist cleaned DI fields by row id for prompt tuning datasets."""

    def fake_compress(pdf_path: Path, *, dest_dir: Path, dpi: int, name_suffix: str) -> Path:
        """Stub compression helper returning a deterministic archive path."""

        dest_dir.mkdir(parents=True, exist_ok=True)
        archived = dest_dir / f"{pdf_path.stem}_{name_suffix}.pdf"
        archived.write_bytes(b"%PDF-1.4\narchived\n%%EOF")
        return archived

    def fake_analyze(pdf_path: Path, *, model_id: str, return_fields: bool) -> Any:
        """Stub extraction helper returning office fields payload."""

        payload = {
            "brutto": 123.45,
            "netto": 100.0,
            "invoice_id": "INV-001",
            "confidence_brutto": 0.95,
            "confidence_netto": 0.91,
            "confidence_invoice_id": 0.87,
        }
        office_fields = {"raw_invoice_no": "INV-RAW-001"}
        if return_fields:
            return payload, office_fields
        return payload

    def fake_clean(fields_payload: dict[str, Any]) -> dict[str, Any]:
        """Stub clean adapter returning deterministic distilled DI dict."""

        return {"invoice_no": fields_payload.get("raw_invoice_no")}

    def fake_semantics(distilled_fields: dict[str, Any]) -> dict[str, Any]:
        """Stub semantic adapter returning office type/sender info."""

        return {
            "purpose": "Service&Andere",
            "sender": "Vendor A",
            "receiver": "Ramen Ippin Dortmund GmbH",
            "receiver_address": "Reinoldistr.8 44135 Dortmund",
            "distilled_echo": distilled_fields,
        }

    monkeypatch.setattr("bills_analysis.integrations.local_backend._compress_pdf_for_archive", fake_compress)
    monkeypatch.setattr("bills_analysis.integrations.local_backend._analyze_pdf_with_azure", fake_analyze)
    monkeypatch.setattr("bills_analysis.integrations.local_backend._clean_invoice_fields", fake_clean)
    monkeypatch.setattr("bills_analysis.integrations.local_backend._extract_office_semantics", fake_semantics)

    test_root = Path("outputs") / "pytest_tmp" / str(uuid4())
    test_root.mkdir(parents=True, exist_ok=True)
    src_pdf = test_root / "office_input.pdf"
    src_pdf.write_bytes(b"%PDF-1.4\nsource\n%%EOF")
    req = CreateBatchRequest(
        type="office",
        run_date="04/02/2026",
        inputs=[{"path": str(src_pdf), "category": "office"}],
        metadata={},
    )
    batch = BatchRecord.new(req)
    backend = LocalPipelineBackend(root=test_root / "out")

    asyncio.run(backend.process_batch(batch))

    di_fields_path = backend.root / batch.batch_id / "di_fields" / "row-0001.json"
    assert di_fields_path.exists()
    saved = json.loads(di_fields_path.read_text(encoding="utf-8"))
    assert saved["invoice_no"] == "INV-RAW-001"


def test_local_backend_process_batch_tracks_mixed_result_and_summary(monkeypatch: pytest.MonkeyPatch) -> None:
    """Backend should keep mixed success/failure rows and return processing summary."""

    def fake_compress(pdf_path: Path, *, dest_dir: Path, dpi: int, name_suffix: str) -> Path:
        """Stub compression helper returning an archive path.""" 

        dest_dir.mkdir(parents=True, exist_ok=True)
        archived = dest_dir / f"{pdf_path.stem}_{name_suffix}.pdf"
        archived.write_bytes(b"%PDF-1.4\narchived\n%%EOF")
        return archived

    def fake_analyze(pdf_path: Path, *, model_id: str, return_fields: bool) -> Any:
        """Stub extraction helper that fails for one file and succeeds for another.""" 

        if pdf_path.name == "bad.pdf":
            raise RuntimeError("simulated azure failure")
        return {
            "store_name": "Demo",
            "brutto": "1.00",
            "netto": "0.80",
            "total_tax": "0.20",
            "confidence_store_name": 0.9,
            "confidence_brutto": 0.9,
            "confidence_netto": 0.9,
            "confidence_total_tax": 0.9,
        }

    monkeypatch.setattr("bills_analysis.integrations.local_backend._compress_pdf_for_archive", fake_compress)
    monkeypatch.setattr("bills_analysis.integrations.local_backend._analyze_pdf_with_azure", fake_analyze)

    test_root = Path("outputs") / "pytest_tmp" / str(uuid4())
    test_root.mkdir(parents=True, exist_ok=True)
    bad_pdf = test_root / "bad.pdf"
    bad_pdf.write_bytes(b"%PDF-1.4\nbad\n%%EOF")
    good_pdf = test_root / "good.pdf"
    good_pdf.write_bytes(b"%PDF-1.4\ngood\n%%EOF")
    req = CreateBatchRequest(
        type="daily",
        run_date="04/02/2026",
        inputs=[
            {"path": str(bad_pdf), "category": "bar"},
            {"path": str(good_pdf), "category": "zbon"},
        ],
        metadata={},
    )
    batch = BatchRecord.new(req)
    backend = LocalPipelineBackend(root=test_root / "out")
    done_events: list[dict[str, Any]] = []

    output = asyncio.run(
        backend.process_batch(
            batch,
            on_file_done=lambda event: _append_file_done_event(done_events, event),
        )
    )
    assert output["processing_summary"]["total_count"] == 2
    assert output["processing_summary"]["extracted_count"] == 1
    assert output["processing_summary"]["failed_count"] == 1

    rows = output["review_rows"]
    assert len(rows) == 2
    by_name = {row["filename"]: row for row in rows}
    assert by_name["bad.pdf"]["result"]["run_date"] == "04/02/2026"
    assert by_name["good.pdf"]["result"]["brutto"] == "1.00"
    assert {event["status"] for event in done_events} == {"failed", "extracted"}


def test_local_backend_skips_over_max_pages_and_keeps_empty_review_row(monkeypatch: pytest.MonkeyPatch) -> None:
    """Files over max_pages should skip Azure extraction and still emit review row with skip_reason."""

    class _DocStub:
        """Minimal fitz document stub exposing page_count and context manager protocol."""

        page_count = 6

        def __enter__(self) -> "_DocStub":
            """Return self for context manager enter."""

            return self

        def __exit__(self, exc_type: Any, exc: Any, tb: Any) -> None:
            """No-op context manager exit for fitz open stub."""

            return None

    def fake_compress(pdf_path: Path, *, dest_dir: Path, dpi: int, name_suffix: str) -> Path:
        """Stub compression helper returning deterministic archive path."""

        dest_dir.mkdir(parents=True, exist_ok=True)
        archived = dest_dir / f"{pdf_path.stem}_{name_suffix}.pdf"
        archived.write_bytes(b"%PDF-1.4\narchived\n%%EOF")
        return archived

    def fake_analyze(pdf_path: Path, *, model_id: str, return_fields: bool) -> Any:
        """Fail hard if Azure analyze is called for over-page-limit files."""

        raise AssertionError("azure analyze should be skipped when page count exceeds max_pages")

    monkeypatch.setattr("bills_analysis.integrations.local_backend._compress_pdf_for_archive", fake_compress)
    monkeypatch.setattr("bills_analysis.integrations.local_backend._analyze_pdf_with_azure", fake_analyze)
    monkeypatch.setattr("fitz.open", lambda path: _DocStub())

    test_root = Path("outputs") / "pytest_tmp" / str(uuid4())
    test_root.mkdir(parents=True, exist_ok=True)
    src_pdf = test_root / "too_many_pages.pdf"
    src_pdf.write_bytes(b"%PDF-1.4\nsource\n%%EOF\n" + (b"0" * 2048))
    req = CreateBatchRequest(
        type="daily",
        run_date="04/02/2026",
        inputs=[{"path": str(src_pdf), "category": "zbon"}],
        metadata={},
    )
    batch = BatchRecord.new(req)
    backend = LocalPipelineBackend(root=test_root / "out")

    output = asyncio.run(backend.process_batch(batch))
    assert len(output["review_rows"]) == 1
    row = output["review_rows"][0]
    assert row["filename"] == "too_many_pages.pdf"
    assert row["result"] == {"run_date": "04/02/2026"}
    assert isinstance(row.get("skip_reason"), str)
    assert "page_count=6" in row["skip_reason"]
    assert "max_pages=4" in row["skip_reason"]


def test_local_backend_limits_extract_concurrency_via_env(monkeypatch: pytest.MonkeyPatch) -> None:
    """Extraction branch should honor configured global concurrency limit across files."""

    active_extract = 0
    max_active_extract = 0

    def fake_page_count(pdf_path: Path) -> int | None:
        """Always return valid in-range page count so extraction is scheduled."""

        return 1

    def fake_compress(pdf_path: Path, *, dest_dir: Path, dpi: int, name_suffix: str) -> Path:
        """Stub compression helper returning deterministic archive path."""

        dest_dir.mkdir(parents=True, exist_ok=True)
        archived = dest_dir / f"{pdf_path.stem}_{name_suffix}.pdf"
        archived.write_bytes(b"%PDF-1.4\narchived\n%%EOF")
        return archived

    def fake_analyze(pdf_path: Path, *, model_id: str, return_fields: bool) -> Any:
        """Track extraction overlap and return deterministic payload."""

        nonlocal active_extract, max_active_extract
        active_extract += 1
        max_active_extract = max(max_active_extract, active_extract)
        try:
            import time

            time.sleep(0.12)
            return {
                "store_name": "Demo",
                "brutto": 1.23,
                "netto": 1.0,
                "total_tax": 0.23,
                "confidence_store_name": 0.9,
                "confidence_brutto": 0.9,
                "confidence_netto": 0.9,
                "confidence_total_tax": 0.9,
            }
        finally:
            active_extract -= 1

    monkeypatch.setenv("BACKEND_EXTRACT_CONCURRENCY", "1")
    monkeypatch.setattr("bills_analysis.integrations.local_backend._safe_pdf_page_count", fake_page_count)
    monkeypatch.setattr("bills_analysis.integrations.local_backend._compress_pdf_for_archive", fake_compress)
    monkeypatch.setattr("bills_analysis.integrations.local_backend._analyze_pdf_with_azure", fake_analyze)

    test_root = Path("outputs") / "pytest_tmp" / str(uuid4())
    test_root.mkdir(parents=True, exist_ok=True)
    pdfs = []
    for idx in range(3):
        path = test_root / f"file_{idx + 1}.pdf"
        path.write_bytes(b"%PDF-1.4\nsource\n%%EOF")
        pdfs.append(path)

    req = CreateBatchRequest(
        type="daily",
        run_date="04/02/2026",
        inputs=[{"path": str(path), "category": "zbon"} for path in pdfs],
        metadata={},
    )
    batch = BatchRecord.new(req)
    backend = LocalPipelineBackend(root=test_root / "out")

    output = asyncio.run(backend.process_batch(batch))

    assert len(output["review_rows"]) == 3
    assert output["processing_summary"]["extracted_count"] == 3
    assert max_active_extract == 1


def test_local_backend_runs_extract_and_compress_in_parallel(monkeypatch: pytest.MonkeyPatch) -> None:
    """Per-file compression and extraction should overlap when page count allows extraction."""

    def fake_page_count(pdf_path: Path) -> int | None:
        """Always return valid in-range page count so extraction is scheduled."""

        return 1

    def fake_compress(pdf_path: Path, *, dest_dir: Path, dpi: int, name_suffix: str) -> Path:
        """Slow compression stub used to assert branch overlap."""

        import time

        time.sleep(0.20)
        dest_dir.mkdir(parents=True, exist_ok=True)
        archived = dest_dir / f"{pdf_path.stem}_{name_suffix}.pdf"
        archived.write_bytes(b"%PDF-1.4\narchived\n%%EOF")
        return archived

    def fake_analyze(pdf_path: Path, *, model_id: str, return_fields: bool) -> Any:
        """Slow extraction stub used to assert branch overlap."""

        import time

        time.sleep(0.20)
        return {
            "store_name": "Demo",
            "brutto": 9.99,
            "netto": 8.0,
            "total_tax": 1.99,
            "confidence_store_name": 0.9,
            "confidence_brutto": 0.9,
            "confidence_netto": 0.9,
            "confidence_total_tax": 0.9,
        }

    monkeypatch.setattr("bills_analysis.integrations.local_backend._safe_pdf_page_count", fake_page_count)
    monkeypatch.setattr("bills_analysis.integrations.local_backend._compress_pdf_for_archive", fake_compress)
    monkeypatch.setattr("bills_analysis.integrations.local_backend._analyze_pdf_with_azure", fake_analyze)

    test_root = Path("outputs") / "pytest_tmp" / str(uuid4())
    test_root.mkdir(parents=True, exist_ok=True)
    src_pdf = test_root / "input.pdf"
    src_pdf.write_bytes(b"%PDF-1.4\nsource\n%%EOF\n" + (b"0" * (2 * 1024 * 1024)))
    req = CreateBatchRequest(
        type="daily",
        run_date="04/02/2026",
        inputs=[{"path": str(src_pdf), "category": "zbon"}],
        metadata={},
    )
    batch = BatchRecord.new(req)
    backend = LocalPipelineBackend(root=test_root / "out")

    import time

    started_at = time.perf_counter()
    output = asyncio.run(backend.process_batch(batch))
    elapsed = time.perf_counter() - started_at

    assert len(output["review_rows"]) == 1
    assert output["processing_summary"]["extracted_count"] == 1
    assert elapsed < 0.34


def test_local_backend_does_not_fail_file_by_timeout_env(monkeypatch: pytest.MonkeyPatch) -> None:
    """File processing should not be forcibly failed by BACKEND_FILE_TIMEOUT_SEC in queueing scenarios."""

    def fake_page_count(pdf_path: Path) -> int | None:
        """Always return valid in-range page count so extraction is scheduled."""

        return 1

    def fake_compress(pdf_path: Path, *, dest_dir: Path, dpi: int, name_suffix: str) -> Path:
        """Slow compression to simulate queueing and long processing."""

        import time

        time.sleep(0.15)
        dest_dir.mkdir(parents=True, exist_ok=True)
        archived = dest_dir / f"{pdf_path.stem}_{name_suffix}.pdf"
        archived.write_bytes(b"%PDF-1.4\narchived\n%%EOF")
        return archived

    def fake_analyze(pdf_path: Path, *, model_id: str, return_fields: bool) -> Any:
        """Slow extract stub that still returns successful payload."""

        import time

        time.sleep(0.15)
        return {
            "store_name": "Demo",
            "brutto": 7.77,
            "netto": 6.0,
            "total_tax": 1.77,
            "confidence_store_name": 0.9,
            "confidence_brutto": 0.9,
            "confidence_netto": 0.9,
            "confidence_total_tax": 0.9,
        }

    monkeypatch.setenv("BACKEND_FILE_TIMEOUT_SEC", "0.01")
    monkeypatch.setattr("bills_analysis.integrations.local_backend._safe_pdf_page_count", fake_page_count)
    monkeypatch.setattr("bills_analysis.integrations.local_backend._compress_pdf_for_archive", fake_compress)
    monkeypatch.setattr("bills_analysis.integrations.local_backend._analyze_pdf_with_azure", fake_analyze)

    test_root = Path("outputs") / "pytest_tmp" / str(uuid4())
    test_root.mkdir(parents=True, exist_ok=True)
    src_pdf = test_root / "slow.pdf"
    src_pdf.write_bytes(b"%PDF-1.4\nsource\n%%EOF")
    req = CreateBatchRequest(
        type="daily",
        run_date="04/02/2026",
        inputs=[{"path": str(src_pdf), "category": "zbon"}],
        metadata={},
    )
    batch = BatchRecord.new(req)
    backend = LocalPipelineBackend(root=test_root / "out")

    output = asyncio.run(backend.process_batch(batch))
    assert output["processing_summary"]["extracted_count"] == 1
    row = output["review_rows"][0]
    assert row["filename"] == "slow.pdf"
    assert row["result"]["brutto"] == 7.77


def test_local_backend_marks_skip_compress_and_copies_archive(monkeypatch: pytest.MonkeyPatch) -> None:
    """Small files under compression-skip threshold should bypass compression and use direct archive copy."""

    called: dict[str, int] = {"compress": 0, "analyze": 0}

    def fake_page_count(pdf_path: Path) -> int | None:
        """Keep page count in-range so DI is still executed."""

        return 1

    def fake_compress(pdf_path: Path, *, dest_dir: Path, dpi: int, name_suffix: str) -> Path:
        """Compression should not be called when skip-compress marker is active."""

        called["compress"] += 1
        raise AssertionError("compression should be skipped for under-threshold files")

    def fake_analyze(pdf_path: Path, *, model_id: str, return_fields: bool) -> Any:
        """Return deterministic extraction payload for non-skipped DI path."""

        called["analyze"] += 1
        return {
            "store_name": "Demo",
            "brutto": 10.0,
            "netto": 8.0,
            "total_tax": 2.0,
            "confidence_store_name": 0.9,
            "confidence_brutto": 0.9,
            "confidence_netto": 0.9,
            "confidence_total_tax": 0.9,
        }

    monkeypatch.setenv("BACKEND_COMPRESS_SKIP_MB", "1")
    monkeypatch.setattr("bills_analysis.integrations.local_backend._safe_pdf_page_count", fake_page_count)
    monkeypatch.setattr("bills_analysis.integrations.local_backend._compress_pdf_for_archive", fake_compress)
    monkeypatch.setattr("bills_analysis.integrations.local_backend._analyze_pdf_with_azure", fake_analyze)

    test_root = Path("outputs") / "pytest_tmp" / str(uuid4())
    test_root.mkdir(parents=True, exist_ok=True)
    src_pdf = test_root / "small_input.pdf"
    src_pdf.write_bytes(b"%PDF-1.4\nsource\n%%EOF\n" + (b"0" * 2048))
    req = CreateBatchRequest(
        type="daily",
        run_date="04/02/2026",
        inputs=[{"path": str(src_pdf), "category": "zbon"}],
        metadata={},
    )
    batch = BatchRecord.new(req)
    backend = LocalPipelineBackend(root=test_root / "out")

    output = asyncio.run(backend.process_batch(batch))
    row = output["review_rows"][0]
    results_payload = json.loads(Path(output["artifacts"]["result_json_path"]).read_text(encoding="utf-8"))
    stored_row = results_payload["items"][0]

    assert called["compress"] == 0
    assert called["analyze"] == 1
    assert output["processing_summary"]["extracted_count"] == 1
    assert row["skip_reason"] is None
    assert Path(str(row["preview_path"])).exists()
    assert "skip-compress" in list(stored_row.get("markers") or [])


def test_local_backend_marks_dual_skip_di_and_skip_compress(monkeypatch: pytest.MonkeyPatch) -> None:
    """Files can carry both skipped-di and skip-compress markers with DI skipped and archive preserved."""

    called: dict[str, int] = {"compress": 0, "analyze": 0}

    def fake_page_count(pdf_path: Path) -> int | None:
        """Return over-limit page count so DI path is skipped."""

        return 7

    def fake_compress(pdf_path: Path, *, dest_dir: Path, dpi: int, name_suffix: str) -> Path:
        """Compression should not run when skip-compress is active."""

        called["compress"] += 1
        raise AssertionError("compression should be skipped when skip-compress marker exists")

    def fake_analyze(pdf_path: Path, *, model_id: str, return_fields: bool) -> Any:
        """DI should be skipped when skipped-di marker is active."""

        called["analyze"] += 1
        raise AssertionError("azure analyze should be skipped when page count exceeds max_pages")

    monkeypatch.setenv("BACKEND_COMPRESS_SKIP_MB", "1")
    monkeypatch.setattr("bills_analysis.integrations.local_backend._safe_pdf_page_count", fake_page_count)
    monkeypatch.setattr("bills_analysis.integrations.local_backend._compress_pdf_for_archive", fake_compress)
    monkeypatch.setattr("bills_analysis.integrations.local_backend._analyze_pdf_with_azure", fake_analyze)

    test_root = Path("outputs") / "pytest_tmp" / str(uuid4())
    test_root.mkdir(parents=True, exist_ok=True)
    src_pdf = test_root / "large_over_pages.pdf"
    src_pdf.write_bytes(b"%PDF-1.4\nsource\n%%EOF\n" + (b"0" * (2 * 1024 * 1024)))
    req = CreateBatchRequest(
        type="daily",
        run_date="04/02/2026",
        inputs=[{"path": str(src_pdf), "category": "zbon"}],
        metadata={},
    )
    batch = BatchRecord.new(req)
    backend = LocalPipelineBackend(root=test_root / "out")

    output = asyncio.run(backend.process_batch(batch))
    row = output["review_rows"][0]
    results_payload = json.loads(Path(output["artifacts"]["result_json_path"]).read_text(encoding="utf-8"))
    stored_row = results_payload["items"][0]

    assert called["compress"] == 0
    assert called["analyze"] == 0
    assert row["skip_reason"] is not None
    assert "max_pages=4" in str(row["skip_reason"])
    assert Path(str(row["preview_path"])).exists()
    markers = list(stored_row.get("markers") or [])
    assert "skipped-di" in markers
    assert "skip-compress" in markers


def test_local_backend_uses_compressed_pdf_as_di_input_for_very_large_files(monkeypatch: pytest.MonkeyPatch) -> None:
    """Files over DI-compressed threshold should run DI on compressed artifact path."""

    called: dict[str, Any] = {"compress": 0, "analyze_paths": []}

    def fake_page_count(pdf_path: Path) -> int | None:
        """Keep page count in-range so extraction runs."""

        return 1

    def fake_compress(pdf_path: Path, *, dest_dir: Path, dpi: int, name_suffix: str) -> Path:
        """Produce deterministic compressed archive artifact path."""

        called["compress"] += 1
        dest_dir.mkdir(parents=True, exist_ok=True)
        archived = dest_dir / f"{pdf_path.stem}_{name_suffix}.pdf"
        archived.write_bytes(b"%PDF-1.4\ncompressed\n%%EOF")
        return archived

    def fake_analyze(pdf_path: Path, *, model_id: str, return_fields: bool) -> Any:
        """Capture DI input path and return deterministic extraction payload."""

        called["analyze_paths"].append(str(pdf_path))
        return {
            "store_name": "Demo",
            "brutto": 11.0,
            "netto": 9.0,
            "total_tax": 2.0,
            "confidence_store_name": 0.9,
            "confidence_brutto": 0.9,
            "confidence_netto": 0.9,
            "confidence_total_tax": 0.9,
        }

    monkeypatch.setenv("BACKEND_COMPRESS_SKIP_MB", "1")
    monkeypatch.setenv("BACKEND_DI_COMPRESSED_INPUT_MB", "3")
    monkeypatch.setattr("bills_analysis.integrations.local_backend._safe_pdf_page_count", fake_page_count)
    monkeypatch.setattr("bills_analysis.integrations.local_backend._compress_pdf_for_archive", fake_compress)
    monkeypatch.setattr("bills_analysis.integrations.local_backend._analyze_pdf_with_azure", fake_analyze)

    test_root = Path("outputs") / "pytest_tmp" / str(uuid4())
    test_root.mkdir(parents=True, exist_ok=True)
    src_pdf = test_root / "very_large.pdf"
    src_pdf.write_bytes(b"%PDF-1.4\nsource\n%%EOF\n" + (b"0" * (3 * 1024 * 1024 + 4096)))
    req = CreateBatchRequest(
        type="daily",
        run_date="04/02/2026",
        inputs=[{"path": str(src_pdf), "category": "zbon"}],
        metadata={},
    )
    batch = BatchRecord.new(req)
    backend = LocalPipelineBackend(root=test_root / "out")

    output = asyncio.run(backend.process_batch(batch))
    assert called["compress"] == 1
    assert len(called["analyze_paths"]) == 1
    analyzed = Path(called["analyze_paths"][0])
    assert analyzed != src_pdf
    assert "archive" in str(analyzed).lower()
    assert output["processing_summary"]["extracted_count"] == 1


def test_local_backend_merge_builds_non_empty_daily_validated_excel() -> None:
    """Daily merge should build non-empty validated workbook from saved review results."""

    test_root = Path("outputs") / "pytest_tmp" / str(uuid4())
    test_root.mkdir(parents=True, exist_ok=True)

    monthly_path = test_root / "monthly.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["Datum", "Umsatz Brutto", "Umsatz Netto"])
    ws.append(["04/02/2026", 0, 0])
    wb.save(monthly_path)

    req = CreateBatchRequest(
        type="daily",
        run_date="04/02/2026",
        inputs=[{"path": str(test_root / "dummy.pdf"), "category": "zbon"}],
        metadata={},
    )
    batch = BatchRecord.new(req)
    batch.review_rows = [
        {
            "row_id": "row-0001",
            "category": "zbon",
            "filename": "zbon.pdf",
            "result": {"run_date": "04/02/2026", "brutto": "123.45", "netto": "100.00"},
            "score": {"brutto": 0.95, "netto": 0.95},
            "preview_path": str(test_root / "preview.pdf"),
        },
        {
            "row_id": "row-0002",
            "category": "bar",
            "filename": "bar.pdf",
            "result": {
                "run_date": "04/02/2026",
                "store_name": "Demo Store",
                "brutto": "23.45",
                "netto": "20.00",
            },
            "score": {"store_name": 0.9, "brutto": 0.9, "netto": 0.9},
            "preview_path": str(test_root / "preview.pdf"),
        },
    ]

    backend = LocalPipelineBackend(root=test_root / "out")
    output = asyncio.run(
        backend.merge_batch(
            batch,
            {"mode": "overwrite", "monthly_excel_path": str(monthly_path)},
        )
    )
    validated_excel = Path(output["validated_excel_path"])
    assert validated_excel.exists()

    merged_wb = load_workbook(validated_excel)
    merged_ws = merged_wb.active
    headers = [cell.value for cell in merged_ws[1]]
    assert merged_ws.max_row >= 2
    assert "Umsatz Brutto" in headers
    brutto_col = headers.index("Umsatz Brutto") + 1
    assert merged_ws.cell(row=2, column=brutto_col).value is not None


def test_local_backend_daily_merge_creates_missing_monthly_and_supports_append() -> None:
    """Daily merge should auto-create missing monthly workbook and honor append mode."""

    test_root = Path("outputs") / "pytest_tmp" / str(uuid4())
    test_root.mkdir(parents=True, exist_ok=True)
    missing_monthly = test_root / "missing_daily_monthly.xlsx"

    req = CreateBatchRequest(
        type="daily",
        run_date="04/02/2026",
        inputs=[{"path": str(test_root / "dummy.pdf"), "category": "zbon"}],
        metadata={},
    )
    batch = BatchRecord.new(req)
    batch.review_rows = [
        {
            "row_id": "row-0001",
            "category": "zbon",
            "filename": "zbon.pdf",
            "result": {"run_date": "04/02/2026", "brutto": "123.45", "netto": "100.00"},
            "score": {"brutto": 0.95, "netto": 0.95},
            "preview_path": str(test_root / "preview.pdf"),
        }
    ]
    backend = LocalPipelineBackend(root=test_root / "out")

    first_output = asyncio.run(
        backend.merge_batch(
            batch,
            {"mode": "overwrite", "monthly_excel_path": str(missing_monthly)},
        )
    )
    assert missing_monthly.exists()
    first_merged = load_workbook(Path(first_output["merged_excel_abs_path"])).active
    assert first_merged.max_row == 2

    second_output = asyncio.run(
        backend.merge_batch(
            batch,
            {"mode": "append", "monthly_excel_path": first_output["merged_excel_abs_path"]},
        )
    )
    second_merged = load_workbook(Path(second_output["merged_excel_abs_path"])).active
    assert second_merged.max_row == 3
    assert second_merged.cell(row=2, column=1).value.strftime("%d/%m/%Y") == "04/02/2026"
    assert second_merged.cell(row=3, column=1).value.strftime("%d/%m/%Y") == "04/02/2026"


def test_local_backend_office_merge_auto_creates_monthly_template_and_supports_append() -> None:
    """Office merge should auto-create missing monthly workbook and honor append mode."""

    test_root = Path("outputs") / "pytest_tmp" / str(uuid4())
    test_root.mkdir(parents=True, exist_ok=True)
    req = CreateBatchRequest(
        type="office",
        run_date="04/02/2026",
        inputs=[{"path": str(test_root / "dummy.pdf"), "category": "office"}],
        metadata={},
    )
    batch = BatchRecord.new(req)
    batch.review_rows = [
        {
            "row_id": "row-0001",
            "category": "office",
            "filename": "office.pdf",
            "result": {
                "run_date": "04/02/2026",
                "type": "Miete",
                "sender": "Metro",
                "brutto": "123.45",
                "netto": "100.00",
                "tax_id": "DE123",
                "receiver_ok": True,
            },
            "score": {"brutto": 0.95, "netto": 0.95},
            "preview_path": str(test_root / "preview.pdf"),
        }
    ]
    backend = LocalPipelineBackend(root=test_root / "out")
    auto_monthly = backend.root / batch.batch_id / "merge_source" / "auto_office_monthly.xlsx"

    first_output = asyncio.run(
        backend.merge_batch(
            batch,
            {"mode": "overwrite"},
        )
    )
    assert auto_monthly.exists()
    validated_ws = load_workbook(Path(first_output["validated_excel_path"])).active
    validated_headers = [cell.value for cell in validated_ws[1]]
    assert "Is Receiver Address OK" not in validated_headers
    first_merged = load_workbook(Path(first_output["merged_excel_abs_path"])).active
    merged_headers = [cell.value for cell in first_merged[1]]
    assert "Is Receiver Address OK" not in merged_headers
    assert first_merged.max_row == 2

    second_output = asyncio.run(
        backend.merge_batch(
            batch,
            {"mode": "append", "monthly_excel_path": first_output["merged_excel_abs_path"]},
        )
    )
    second_merged = load_workbook(Path(second_output["merged_excel_abs_path"])).active
    assert second_merged.max_row == 3
    assert second_merged.cell(row=2, column=1).value.strftime("%d/%m/%Y") == "04/02/2026"
    assert second_merged.cell(row=3, column=1).value.strftime("%d/%m/%Y") == "04/02/2026"


def test_local_backend_office_overwrite_does_not_collapse_multiple_same_datum_rows() -> None:
    """Office overwrite should retain multiple reviewed rows sharing the same Datum."""

    test_root = Path("outputs") / "pytest_tmp" / str(uuid4())
    test_root.mkdir(parents=True, exist_ok=True)

    req = CreateBatchRequest(
        type="office",
        run_date="15/02/2026",
        inputs=[{"path": str(test_root / "dummy_1.pdf"), "category": "office"}],
        metadata={},
    )
    batch = BatchRecord.new(req)
    batch.review_rows = [
        {
            "row_id": "row-0001",
            "category": "office",
            "filename": "a.pdf",
            "result": {"run_date": "15/02/2026", "type": "A", "sender": "Vendor-A", "receiver_ok": True},
            "score": {},
        },
        {
            "row_id": "row-0002",
            "category": "office",
            "filename": "b.pdf",
            "result": {"run_date": "15/02/2026", "type": "B", "sender": "Vendor-B", "receiver_ok": False},
            "score": {},
        },
        {
            "row_id": "row-0003",
            "category": "office",
            "filename": "c.pdf",
            "result": {"run_date": "15/02/2026", "type": "C", "sender": "Vendor-C", "receiver_ok": True},
            "score": {},
        },
    ]
    backend = LocalPipelineBackend(root=test_root / "out")

    output = asyncio.run(backend.merge_batch(batch, {"mode": "overwrite"}))
    merged_ws = load_workbook(Path(output["merged_excel_abs_path"])).active
    assert merged_ws.max_row == 4
    merged_senders = [merged_ws.cell(row=row_idx, column=3).value for row_idx in range(2, 5)]
    assert merged_senders == ["Vendor-A", "Vendor-B", "Vendor-C"]


def _openapi_contract_subset(spec: dict) -> dict:
    """Extract only v1 paths and schema components for snapshot comparison."""

    paths = {}
    for path, methods in spec.get("paths", {}).items():
        if not path.startswith("/v1/batches") and path != "/healthz":
            continue
        paths[path] = methods
    components = spec.get("components", {}).get("schemas", {})
    return {"paths": paths, "schemas": components}


def test_openapi_contract_frozen_v1() -> None:
    """Current OpenAPI subset must match frozen baseline."""

    _, app = _get_test_client_and_app()
    baseline_path = Path("tests/openapi_v1_baseline.json")
    if not baseline_path.exists():
        pytest.skip(
            "Missing openapi baseline. Run: PYTHONPATH=src python scripts/export_openapi_v1.py"
        )
    baseline = json.loads(baseline_path.read_text(encoding="utf-8"))
    current = _openapi_contract_subset(app.openapi())
    assert current == baseline
