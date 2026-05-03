from __future__ import annotations

import json
from pathlib import Path
from uuid import uuid4

from openpyxl import load_workbook

from bills_analysis.services.review_service import (
    export_daily_review_excel,
    export_office_review_excel,
)


def _write_json(path: Path, payload: list[dict]) -> None:
    """Write list payload into JSON file for mapping tests."""

    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def test_daily_excel_mapping_parity() -> None:
    """Daily mapping should preserve row structure and output key business columns."""

    root = Path("outputs") / "pytest_tmp" / str(uuid4())
    json_path = root / "daily.json"
    _write_json(
        json_path,
        [
            {
                "filename": "zbon.pdf",
                "category": "zbon",
                "result": {"run_date": "04/02/2026", "brutto": "100.00", "netto": "80.00"},
                "score": {"brutto": 0.9, "netto": 0.9},
            },
            {
                "filename": "bar.pdf",
                "category": "bar",
                "result": {
                    "run_date": "04/02/2026",
                    "store_name": "REWE",
                    "brutto": "20.00",
                    "netto": "16.00",
                    "bill_id": "RE-1001",
                },
                "score": {"store_name": 0.9, "brutto": 0.9, "netto": 0.9},
                "preview_path": str(root / "bar.pdf"),
            },
        ],
    )

    out_path = export_daily_review_excel(json_path, config_path=Path("tests/config.json"))
    wb = load_workbook(out_path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    values = [cell.value for cell in ws[2]]
    row = {headers[idx]: values[idx] for idx in range(len(headers))}

    assert row["Datum"] == "04/02/2026"
    assert float(row["Umsatz Brutto"]) == 100.0
    assert row["Ausgabe 1 Name"] == "REWE"
    assert row["Ausgabe 1 Rechnung-Nr"] == "RE-1001"
    assert float(row["Ausgabe 1 Brutto"]) == 20.0


def test_office_excel_mapping_parity() -> None:
    """Office mapping should keep review columns and need-review marking behavior."""

    root = Path("outputs") / "pytest_tmp" / str(uuid4())
    json_path = root / "office.json"
    _write_json(
        json_path,
        [
            {
                "filename": "office.pdf",
                "category": "office",
                "result": {
                    "run_date": "04/02/2026 10:20:30",
                    "type": "Miete",
                    "sender": "Metro",
                    "brutto": "50.00",
                    "netto": "40.00",
                    "tax_id": "",
                    "receiver_ok": False,
                },
                "score": {"brutto": 0.9, "netto": 0.9},
                "preview_path": str(root / "office.pdf"),
            }
        ],
    )

    out_path = export_office_review_excel(json_path, config_path=Path("tests/config.json"))
    wb = load_workbook(out_path)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    values = [cell.value for cell in ws[2]]
    row = {headers[idx]: values[idx] for idx in range(len(headers))}

    assert "Is Receiver OK" in headers
    assert "Is Receiver Address OK" not in headers
    assert row["Datum"].strftime("%d/%m/%Y") == "04/02/2026"
    assert row["Type"] == "Miete"
    assert row["Rechnung Name"] == "Metro"
    assert row["need review"] is True
    scan_col = headers.index("Rechnung Scannen") + 1
    assert ws.cell(row=2, column=scan_col).value == "check pdf"
