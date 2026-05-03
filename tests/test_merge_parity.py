from __future__ import annotations

from pathlib import Path
from uuid import uuid4

from openpyxl import Workbook, load_workbook

from bills_analysis.services.merge_service import merge_daily, merge_office


def _new_book(path: Path, headers: list[str], rows: list[list[object]]) -> None:
    """Create workbook fixture with headers and data rows."""

    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append(row)
    wb.save(path)


def _sheet_rows_as_values(path: Path) -> list[list[object]]:
    """Read worksheet values from row2 onward for deterministic assertions."""

    wb = load_workbook(path)
    ws = wb.active
    return [[ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)] for r in range(2, ws.max_row + 1)]


def test_daily_merge_overwrite_existing_datum() -> None:
    """Daily overwrite mode should update existing matching Datum row."""

    root = Path("outputs") / "pytest_tmp" / str(uuid4())
    validated = root / "validated_daily.xlsx"
    monthly = root / "monthly_daily.xlsx"

    _new_book(
        validated,
        ["Datum", "Umsatz Brutto", "Umsatz Netto", "need review"],
        [["04/02/2026", 120.0, 100.0, False]],
    )
    _new_book(
        monthly,
        ["Datum", "Umsatz Brutto", "Umsatz Netto"],
        [["04/02/2026", 1.0, 2.0]],
    )

    out_path = merge_daily(validated, monthly, out_dir=root)
    wb = load_workbook(out_path)
    ws = wb.active
    assert ws.cell(row=2, column=2).value == 120.0
    assert ws.cell(row=2, column=3).value == 100.0


def test_daily_merge_overwrite_missing_datum_appends() -> None:
    """Daily overwrite mode should append when Datum does not yet exist."""

    root = Path("outputs") / "pytest_tmp" / str(uuid4())
    validated = root / "validated_daily.xlsx"
    monthly = root / "monthly_daily.xlsx"

    _new_book(
        validated,
        ["Datum", "Umsatz Brutto", "Umsatz Netto", "need review"],
        [["05/02/2026", 120.0, 100.0, False]],
    )
    _new_book(
        monthly,
        ["Datum", "Umsatz Brutto", "Umsatz Netto"],
        [["04/02/2026", 1.0, 2.0]],
    )

    out_path = merge_daily(validated, monthly, out_dir=root)
    rows = _sheet_rows_as_values(out_path)
    assert len(rows) == 2
    assert rows[1][0].strftime("%d/%m/%Y") == "05/02/2026"
    assert rows[1][1] == 120.0
    assert rows[1][2] == 100.0


def test_daily_merge_append_allows_duplicate_datum() -> None:
    """Daily append mode should always add a new row even with duplicate Datum."""

    root = Path("outputs") / "pytest_tmp" / str(uuid4())
    validated = root / "validated_daily.xlsx"
    monthly = root / "monthly_daily.xlsx"

    _new_book(
        validated,
        ["Datum", "Umsatz Brutto", "Umsatz Netto", "need review"],
        [["04/02/2026", 120.0, 100.0, False]],
    )
    _new_book(
        monthly,
        ["Datum", "Umsatz Brutto", "Umsatz Netto"],
        [["04/02/2026", 1.0, 2.0]],
    )

    out_path = merge_daily(validated, monthly, out_dir=root, append=True)
    rows = _sheet_rows_as_values(out_path)
    assert len(rows) == 2
    assert rows[0][0].strftime("%d/%m/%Y") == "04/02/2026"
    assert rows[1][0].strftime("%d/%m/%Y") == "04/02/2026"
    assert rows[0][1] == 1.0
    assert rows[1][1] == 120.0


def test_daily_merge_creates_missing_monthly_template() -> None:
    """Daily merge should create a canonical monthly workbook when target file is missing."""

    root = Path("outputs") / "pytest_tmp" / str(uuid4())
    validated = root / "validated_daily.xlsx"
    monthly = root / "missing_monthly_daily.xlsx"

    _new_book(
        validated,
        ["Datum", "Umsatz Brutto", "Umsatz Netto", "need review"],
        [["04/02/2026", 120.0, 100.0, False]],
    )
    assert not monthly.exists()

    out_path = merge_daily(validated, monthly, out_dir=root)
    assert monthly.exists()
    wb_monthly = load_workbook(monthly)
    ws_monthly = wb_monthly.active
    headers = [cell.value for cell in ws_monthly[1]]
    assert headers[:4] == ["Datum", "Umsatz Brutto", "Umsatz Netto", "Wie viel Rechnungen"]
    assert "Ausgabe 1 Rechnung-Nr" in headers
    assert "Ausgabe sum Brutto" in headers
    assert "Ausgabe Sum Netto" in headers

    rows = _sheet_rows_as_values(out_path)
    assert len(rows) == 1
    assert rows[0][1] == 120.0
    assert rows[0][2] == 100.0


def test_daily_merge_writes_expense_sum_columns() -> None:
    """Daily merge should write Brutto and Netto sum columns when present."""

    root = Path("outputs") / "pytest_tmp" / str(uuid4())
    validated = root / "validated_daily.xlsx"
    monthly = root / "monthly_daily.xlsx"

    _new_book(
        validated,
        [
            "Datum",
            "Ausgabe 1 Rechnung-Nr",
            "Ausgabe 1 Brutto",
            "Ausgabe 1 Netto",
            "Ausgabe 2 Rechnung-Nr",
            "Ausgabe 2 Brutto",
            "Ausgabe 2 Netto",
            "need review",
        ],
        [["04/02/2026", "RE-1", 10.0, 8.0, "RE-2", 5.5, 4.5, False]],
    )
    _new_book(
        monthly,
        [
            "Datum",
            "Ausgabe 1 Brutto",
            "Ausgabe 1 Netto",
            "Ausgabe 2 Brutto",
            "Ausgabe 2 Netto",
            "Ausgabe sum Brutto",
            "Ausgabe Sum Netto",
        ],
        [],
    )

    out_path = merge_daily(validated, monthly, out_dir=root)
    wb = load_workbook(out_path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    brutto_col = headers.index("Ausgabe sum Brutto") + 1
    netto_col = headers.index("Ausgabe Sum Netto") + 1
    assert ws.cell(row=2, column=brutto_col).value == 15.5
    assert ws.cell(row=2, column=netto_col).value == 12.5


def test_daily_merge_always_sorts_by_datum_ascending() -> None:
    """Daily merge result should be sorted by Datum ascending after merge."""

    root = Path("outputs") / "pytest_tmp" / str(uuid4())
    validated = root / "validated_daily.xlsx"
    monthly = root / "monthly_daily.xlsx"

    _new_book(
        validated,
        ["Datum", "Umsatz Brutto", "Umsatz Netto", "need review"],
        [["04/02/2026", 120.0, 100.0, False]],
    )
    _new_book(
        monthly,
        ["Datum", "Umsatz Brutto", "Umsatz Netto"],
        [["06/02/2026", 6.0, 6.0], ["05/02/2026", 5.0, 5.0]],
    )

    out_path = merge_daily(validated, monthly, out_dir=root)
    rows = _sheet_rows_as_values(out_path)
    assert [row[0].strftime("%d/%m/%Y") for row in rows] == ["04/02/2026", "05/02/2026", "06/02/2026"]


def test_office_merge_append_parity() -> None:
    """Office merge append mode should append validated rows and keep hyperlinks."""

    root = Path("outputs") / "pytest_tmp" / str(uuid4())
    validated = root / "validated_office.xlsx"
    monthly = root / "monthly_office.xlsx"

    _new_book(
        validated,
        ["Datum", "Type", "Rechnung Name", "need review", "Rechnung Scannen"],
        [["04/02/2026", "Miete", "Metro", True, "check pdf"]],
    )
    wb_validated = load_workbook(validated)
    ws_validated = wb_validated.active
    ws_validated.cell(row=2, column=5).hyperlink = "file:///tmp/demo.pdf"
    wb_validated.save(validated)

    _new_book(
        monthly,
        ["Datum", "Type", "Rechnung Name", "Rechnung Scannen"],
        [["03/02/2026", "Alt", "Old", None]],
    )

    out_path = merge_office(validated, monthly, out_dir=root, append=True)
    wb = load_workbook(out_path)
    ws = wb.active
    assert ws.max_row == 3
    assert ws.cell(row=3, column=2).value == "Miete"
    assert ws.cell(row=3, column=3).value == "Metro"


def test_office_merge_creates_missing_monthly_template() -> None:
    """Office merge should create a canonical monthly workbook when target file is missing."""

    root = Path("outputs") / "pytest_tmp" / str(uuid4())
    validated = root / "validated_office.xlsx"
    monthly = root / "missing_monthly_office.xlsx"

    _new_book(
        validated,
        ["Datum", "Type", "Rechnung Name", "need review", "Rechnung Scannen"],
        [["04/02/2026", "Miete", "Metro", False, "check pdf"]],
    )
    assert not monthly.exists()

    out_path = merge_office(validated, monthly, out_dir=root, append=True)
    assert monthly.exists()
    wb_monthly = load_workbook(monthly)
    ws_monthly = wb_monthly.active
    headers = [cell.value for cell in ws_monthly[1]]
    assert headers == [
        "Datum",
        "Type",
        "Rechnung Name",
        "Brutto",
        "Netto",
        "Steuernummer",
        "Is Receiver OK",
        "Rechnung Scannen",
    ]

    wb_out = load_workbook(out_path)
    ws_out = wb_out.active
    assert ws_out.max_row == 2
    assert ws_out.cell(row=2, column=2).value == "Miete"
    assert ws_out.cell(row=2, column=3).value == "Metro"


def test_office_merge_overwrite_keeps_multiple_same_datum_rows_from_validated() -> None:
    """Office overwrite mode should not collapse multiple validated rows sharing the same Datum."""

    root = Path("outputs") / "pytest_tmp" / str(uuid4())
    validated = root / "validated_office.xlsx"
    monthly = root / "monthly_office.xlsx"

    _new_book(
        validated,
        ["Datum", "Type", "Rechnung Name", "need review", "Rechnung Scannen"],
        [
            ["04/02/2026", "Miete", "Metro", False, "check pdf"],
            ["04/02/2026", "Strom", "EON", False, "check pdf"],
            ["04/02/2026", "Internet", "Telekom", False, "check pdf"],
        ],
    )
    _new_book(
        monthly,
        ["Datum", "Type", "Rechnung Name", "Rechnung Scannen"],
        [],
    )

    out_path = merge_office(validated, monthly, out_dir=root, append=False)
    wb_out = load_workbook(out_path)
    ws_out = wb_out.active
    assert ws_out.max_row == 4
    names = [ws_out.cell(row=row_idx, column=3).value for row_idx in range(2, 5)]
    assert names == ["Metro", "EON", "Telekom"]


def test_office_merge_filters_legacy_receiver_address_ok_column() -> None:
    """Office merge should ignore legacy Is Receiver Address OK column from validated input."""

    root = Path("outputs") / "pytest_tmp" / str(uuid4())
    validated = root / "validated_office_legacy.xlsx"
    monthly = root / "monthly_office.xlsx"

    _new_book(
        validated,
        [
            "Datum",
            "Type",
            "Rechnung Name",
            "Brutto",
            "Netto",
            "Steuernummer",
            "Is Receiver OK",
            "Is Receiver Address OK",
            "need review",
            "Rechnung Scannen",
        ],
        [["04/02/2026", "Miete", "Metro", 10.0, 8.0, "DE123", True, False, False, "check pdf"]],
    )
    _new_book(
        monthly,
        ["Datum", "Type", "Rechnung Name", "Brutto", "Netto", "Steuernummer", "Is Receiver OK", "Rechnung Scannen"],
        [],
    )

    out_path = merge_office(validated, monthly, out_dir=root, append=False)
    wb_out = load_workbook(out_path)
    ws_out = wb_out.active
    headers = [cell.value for cell in ws_out[1]]
    assert "Is Receiver Address OK" not in headers
    assert ws_out.cell(row=2, column=7).value is True
