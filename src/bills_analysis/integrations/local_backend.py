from __future__ import annotations

import asyncio
import json
import logging
import os
import shutil
from datetime import UTC, datetime
from math import inf
from pathlib import Path
from typing import Any, Awaitable, Callable

import fitz
from openpyxl import Workbook

from bills_analysis.excel_ops import normalize_date, write_datum_cell
from bills_analysis.integrations.excel_mapper_adapter import map_daily_json_to_excel
from bills_analysis.integrations.office_receiver_mapping import resolve_expected_receiver_from_metadata
from bills_analysis.integrations.office_semantics import match_receiver_address, resolve_receiver_ok
from bills_analysis.models.common import InputFile
from bills_analysis.models.internal import BatchRecord
from bills_analysis.services.merge_service import merge_daily, merge_office

LOGGER = logging.getLogger(__name__)


def _is_azure_mock_enabled() -> bool:
    """Return whether local backend should bypass real Azure integrations."""

    return os.getenv("AZURE_MOCK", "").strip().lower() in {"1", "true", "yes", "on"}


def _mock_analyze_pdf_with_azure(*, model_id: str, return_fields: bool) -> Any:
    """Return deterministic Azure DI payloads for local E2E smoke mode."""

    if model_id == "prebuilt-invoice":
        invoice_payload = {
            "brutto": 120.5,
            "netto": 100.0,
            "invoice_id": "INV-1001",
            "confidence_brutto": 0.95,
            "confidence_netto": 0.92,
            "confidence_invoice_id": 0.9,
        }
        if return_fields:
            return invoice_payload, {"invoice_id": "INV-1001", "sender": "Vendor GmbH"}
        return invoice_payload

    receipt_payload = {
        "store_name": "Demo Store",
        "brutto": 12.34,
        "netto": 10.0,
        "total_tax": 2.34,
        "confidence_store_name": 0.9,
        "confidence_brutto": 0.95,
        "confidence_netto": 0.9,
        "confidence_total_tax": 0.86,
    }
    if return_fields:
        return receipt_payload, {}
    return receipt_payload


def _mock_extract_office_semantics(distilled_fields: dict[str, Any]) -> dict[str, Any]:
    """Return deterministic Office semantics payload for local smoke mode."""

    return {
        "purpose": "office-cost",
        "sender": str(distilled_fields.get("sender") or "Vendor GmbH"),
        "receiver": "Ramen Ippin Dortmund GmbH",
        "receiver_address": "Reinoldistr.8 44135 Dortmund",
        "receiver_ok": True,
    }


def _compress_pdf_for_archive(
    pdf_path: Path,
    *,
    dest_dir: Path,
    dpi: int,
    name_suffix: str,
) -> Path:
    """Compress and archive one source PDF via preprocess module."""

    from bills_analysis.preprocess import compress_image_only_pdf

    return compress_image_only_pdf(
        pdf_path,
        dest_dir=dest_dir,
        dpi=dpi,
        name_suffix=name_suffix,
    )


def _copy_pdf_for_archive_without_compression(
    pdf_path: Path,
    *,
    dest_dir: Path,
    name_suffix: str,
) -> Path:
    """Copy one source PDF into archive directly when compression is intentionally skipped."""

    dest_dir.mkdir(parents=True, exist_ok=True)
    target_path = dest_dir / f"{pdf_path.stem}_{name_suffix}.pdf"
    if target_path.resolve() == pdf_path.resolve():
        return target_path
    shutil.copy2(pdf_path, target_path)
    return target_path


def _analyze_pdf_with_azure(
    pdf_path: Path,
    *,
    model_id: str,
    return_fields: bool,
) -> Any:
    """Run Azure DI extraction for one PDF via validated pipeline adapter."""

    if _is_azure_mock_enabled():
        return _mock_analyze_pdf_with_azure(model_id=model_id, return_fields=return_fields)

    from bills_analysis.extract_by_azure_api import analyze_document_with_azure

    return analyze_document_with_azure(
        str(pdf_path),
        model_id=model_id,
        return_fields=return_fields,
    )


def _clean_invoice_fields(fields_payload: dict[str, Any]) -> dict[str, Any]:
    """Normalize raw invoice fields before Office semantic enrichment."""

    if _is_azure_mock_enabled():
        return dict(fields_payload)

    from bills_analysis.extract_by_azure_api import clean_invoice_json

    return clean_invoice_json(fields_payload)


def _extract_office_semantics(distilled_fields: dict[str, Any]) -> dict[str, Any]:
    """Extract Office category/sender/receiver semantic fields with AOAI."""

    if _is_azure_mock_enabled():
        return _mock_extract_office_semantics(distilled_fields)

    from bills_analysis.extract_by_azure_api import extract_office_invoice_azure

    return extract_office_invoice_azure(distilled_fields)


def _to_excel_hyperlink(value: Any) -> str | None:
    """Convert local or remote preview values into an Excel hyperlink target."""

    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    if text.startswith("http://") or text.startswith("https://"):
        return text
    path = Path(text)
    if not path.is_absolute():
        path = path.resolve()
    try:
        return path.as_uri()
    except ValueError:
        return str(path)


_RECEIPT_RATIO_THRESHOLD = 2.0


def _safe_pdf_page_count(pdf_path: Path) -> int | None:
    """Read page count defensively for legacy tests and simple precheck callers."""

    try:
        # Avoid PyMuPDF native crashes on malformed tiny placeholder PDFs.
        if pdf_path.stat().st_size < 1024:
            return None
    except Exception:
        return None
    try:
        with fitz.open(pdf_path) as doc:
            return int(doc.page_count)
    except Exception:
        return None


def _safe_pdf_page_info(pdf_path: Path) -> tuple[int | None, float | None]:
    """Read page count and first-page height/width ratio defensively."""

    count = _safe_pdf_page_count(pdf_path)
    if count is None:
        return None, None
    try:
        with fitz.open(pdf_path) as doc:
            if count < 1:
                return count, None
            page = doc.load_page(0)
            rect = page.rect
            width, height = rect.width, rect.height
            if page.rotation in (90, 270):
                width, height = height, width
            ratio = height / width if width > 0 else None
            return count, ratio
    except Exception:
        return count, None


def _read_int_env(name: str, default: int, *, minimum: int | None = None) -> int:
    """Read integer env var defensively and clamp to optional minimum."""

    try:
        value = int(os.getenv(name, str(default)))
    except Exception:
        value = default
    if minimum is not None:
        value = max(minimum, value)
    return value


def _read_float_env(name: str, default: float, *, minimum: float | None = None) -> float:
    """Read float env var defensively and clamp to optional minimum."""

    try:
        value = float(os.getenv(name, str(default)))
    except Exception:
        value = default
    if minimum is not None:
        value = max(minimum, value)
    return value


class _AsyncRateLimiter:
    """Simple coroutine rate limiter enforcing a minimum start interval."""

    def __init__(self, min_interval_sec: float) -> None:
        """Initialize limiter with monotonic-clock spacing interval."""

        self.min_interval_sec = max(0.0, min_interval_sec)
        self._lock = asyncio.Lock()
        self._next_allowed_at = -inf

    async def wait_turn(self) -> None:
        """Wait until this caller can start, preserving call order under lock."""

        if self.min_interval_sec <= 0:
            return
        loop = asyncio.get_running_loop()
        async with self._lock:
            now = loop.time()
            if now < self._next_allowed_at:
                await asyncio.sleep(self._next_allowed_at - now)
            self._next_allowed_at = loop.time() + self.min_interval_sec


def _resolve_receiver_ok(
    office_info: dict[str, Any],
    *,
    batch_metadata: dict[str, Any] | None = None,
) -> bool | None:
    """Normalize Office receiver consistency output to bool using city-resolved expected receiver."""
    expected = resolve_expected_receiver_from_metadata(batch_metadata)
    expected_receiver = expected["receiver_name"].strip()
    expected_address = expected["receiver_address"].strip()
    if not expected_receiver or not expected_address:
        return None
    return resolve_receiver_ok(
        office_info,
        expected_receiver=expected_receiver,
        expected_address=expected_address,
    )


def _resolve_receiver_address_ok(
    office_info: dict[str, Any],
    *,
    batch_metadata: dict[str, Any] | None = None,
) -> bool | None:
    """Normalize Office receiver address consistency output to bool using city-resolved expected address."""

    receiver_address = office_info.get("receiver_address")
    if not isinstance(receiver_address, str):
        return None
    address_text = receiver_address.strip()
    if not address_text:
        return None
    expected = resolve_expected_receiver_from_metadata(batch_metadata)
    expected_address = expected["receiver_address"].strip()
    if not expected_address:
        return None
    return match_receiver_address(address_text, expected_address)


def _copy_pdf_to_organized_dir(
    *,
    compressed_pdf: Path,
    organized_root: Path,
    category: str,
    run_date: str,
    extracted_result: dict[str, Any],
) -> Path:
    """Copy one compressed PDF into legacy-organized directory with legacy naming rules."""

    from bills_analysis.integrations.azure_pipeline_adapter import (
        get_archive_subdir_name,
        get_compressed_pdf_name,
    )

    organized_dir = organized_root / get_archive_subdir_name(run_date, category)
    organized_dir.mkdir(parents=True, exist_ok=True)
    target_name = get_compressed_pdf_name(category, extracted_result, run_date) or compressed_pdf.name
    target_path = organized_dir / target_name
    if target_path.resolve() == compressed_pdf.resolve():
        return target_path
    shutil.copy2(compressed_pdf, target_path)
    return target_path


class LocalPipelineBackend:
    """Local backend adapter that executes preprocess + extraction flow."""

    def __init__(self, *, root: Path | None = None) -> None:
        """Initialize output root and extraction concurrency controls."""

        self.root = (root or (Path("outputs") / "webapp")).resolve()
        self.file_timeout_sec = self._read_optional_file_timeout()
        self.extract_concurrency = _read_int_env("BACKEND_EXTRACT_CONCURRENCY", 4, minimum=1)
        self.extract_min_interval_sec = _read_float_env("BACKEND_EXTRACT_MIN_INTERVAL_SEC", 0.0, minimum=0.0)
        self.compress_skip_mb = _read_float_env("BACKEND_COMPRESS_SKIP_MB", 1.0, minimum=0.1)
        self.di_compressed_input_mb = _read_float_env("BACKEND_DI_COMPRESSED_INPUT_MB", 3.0, minimum=0.1)

    def _read_optional_file_timeout(self) -> float | None:
        """Return legacy file timeout env value; non-positive and missing values disable it."""

        raw = os.getenv("BACKEND_FILE_TIMEOUT_SEC")
        if raw is None:
            return None
        try:
            value = float(raw)
        except (TypeError, ValueError):
            return None
        return value if value > 0 else None

    async def process_batch(
        self,
        batch: BatchRecord,
        *,
        on_file_done: Callable[[dict[str, Any]], Awaitable[None]] | None = None,
    ) -> dict[str, Any]:
        """Process uploaded PDFs with per-file completion callbacks and summary."""

        out_dir = self.root / batch.batch_id
        out_dir.mkdir(parents=True, exist_ok=True)
        archive_root = out_dir / "archive"
        archive_root.mkdir(parents=True, exist_ok=True)
        organized_root = self.root.parent / "organized"
        organized_root.mkdir(parents=True, exist_ok=True)
        now = datetime.now(UTC).isoformat()
        results_path = out_dir / "results.json"
        review_path = out_dir / "review_rows.json"
        max_pages = 4
        config_path = Path("tests") / "config.json"
        try:
            config_payload = json.loads(config_path.read_text(encoding="utf-8"))
            max_pages = int(config_payload.get("max_pages", max_pages))
        except Exception:
            max_pages = 4
        extract_semaphore = asyncio.Semaphore(self.extract_concurrency)
        extract_rate_limiter = _AsyncRateLimiter(self.extract_min_interval_sec)

        tasks = [
            asyncio.create_task(
                self._process_one_file_with_timeout(
                    row_id=f"row-{idx:04d}",
                    batch=batch,
                    item=item,
                    archive_root=archive_root,
                    organized_root=organized_root,
                    max_pages=max_pages,
                    extract_semaphore=extract_semaphore,
                    extract_rate_limiter=extract_rate_limiter,
                )
            )
            for idx, item in enumerate(batch.inputs, start=1)
        ]
        rows: list[dict[str, Any]] = []
        for completed in asyncio.as_completed(tasks):
            row = await completed
            rows.append(row)
            status = self._row_status(row)
            LOGGER.info(
                "file_done row_id=%s filename=%s status=%s markers=%s skip_reason=%r error=%r",
                row.get("row_id"),
                row.get("filename"),
                status,
                row.get("markers"),
                row.get("skip_reason"),
                self._row_error(row),
            )
            if on_file_done is not None:
                await on_file_done(self._build_file_done_event(row))
        rows.sort(key=lambda row: str(row.get("row_id") or ""))

        results_payload = {
            "batch_id": batch.batch_id,
            "batch_type": batch.batch_type.value,
            "run_date": batch.run_date,
            "inputs": [item.model_dump() for item in batch.inputs],
            "items": rows,
            "generated_at": now,
        }
        review_payload = [
            {
                "row_id": row["row_id"],
                "filename": row["filename"],
                "category": row["category"],
                "result": row["result"],
                "score": row["score"],
                "preview_path": row.get("preview_path"),
                "skip_reason": row.get("skip_reason"),
            }
            for row in rows
        ]
        results_path.write_text(
            json.dumps(results_payload, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        review_path.write_text(
            json.dumps(review_payload, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        extracted_count = sum(1 for row in rows if self._row_status(row) == "extracted")
        failed_count = sum(1 for row in rows if self._row_status(row) == "failed")
        return {
            "artifacts": {
                "result_json_path": str(results_path),
                "review_json_path": str(review_path),
                "archive_root": str(archive_root),
                "organized_root": str(organized_root),
            },
            "review_rows": review_payload,
            "processing_summary": {
                "total_count": len(rows),
                "extracted_count": extracted_count,
                "failed_count": failed_count,
            },
        }

    async def _process_one_file_with_timeout(
        self,
        *,
        row_id: str,
        batch: BatchRecord,
        item: InputFile,
        archive_root: Path,
        organized_root: Path,
        max_pages: int,
        extract_semaphore: asyncio.Semaphore,
        extract_rate_limiter: _AsyncRateLimiter,
    ) -> dict[str, Any]:
        """Execute one file processing without file-level timeout short-circuit."""

        return await self._process_one_file_async(
            row_id=row_id,
            batch=batch,
            item=item,
            archive_root=archive_root,
            organized_root=organized_root,
            max_pages=max_pages,
            extract_semaphore=extract_semaphore,
            extract_rate_limiter=extract_rate_limiter,
        )

    async def _run_extract_with_controls(
        self,
        *,
        pdf_path: Path,
        model_id: str,
        return_fields: bool,
        extract_semaphore: asyncio.Semaphore,
        extract_rate_limiter: _AsyncRateLimiter,
    ) -> Any:
        """Run extraction under shared concurrency and rate controls."""

        async with extract_semaphore:
            await extract_rate_limiter.wait_turn()
            return await asyncio.to_thread(
                _analyze_pdf_with_azure,
                pdf_path,
                model_id=model_id,
                return_fields=return_fields,
            )

    async def _process_one_file_async(
        self,
        *,
        row_id: str,
        batch: BatchRecord,
        item: InputFile,
        archive_root: Path,
        organized_root: Path,
        max_pages: int,
        extract_semaphore: asyncio.Semaphore | None = None,
        extract_rate_limiter: _AsyncRateLimiter | None = None,
    ) -> dict[str, Any]:
        """Run page precheck plus parallel compression/extraction for one input file."""

        if extract_semaphore is None or extract_rate_limiter is None:
            return await asyncio.to_thread(
                self._process_one_file,
                row_id=row_id,
                batch=batch,
                item=item,
                archive_root=archive_root,
                organized_root=organized_root,
                max_pages=max_pages,
            )

        source_path = Path(item.path)
        category = (item.category or "office").lower()
        row: dict[str, Any] = {
            "row_id": row_id,
            "filename": source_path.name,
            "category": category,
            "result": {"run_date": batch.run_date},
            "score": {},
        }
        markers: list[str] = []

        if not source_path.exists():
            row["error"] = f"missing input file: {source_path}"
            return row

        page_count, page_ratio = await asyncio.to_thread(_safe_pdf_page_info, source_path)
        should_extract = True
        force_skip_compress_for_over_pages = False
        if page_count is not None and page_count > max_pages:
            markers.append("skipped-di")
            row["skip_reason"] = f"page_count={page_count} > max_pages={max_pages}"
            should_extract = False
            force_skip_compress_for_over_pages = True
        file_size_mb = 0.0
        try:
            file_size_mb = source_path.stat().st_size / (1024 * 1024)
        except Exception:
            file_size_mb = 0.0
        should_skip_compress = force_skip_compress_for_over_pages or (file_size_mb < self.compress_skip_mb)
        use_compressed_for_di = file_size_mb > self.di_compressed_input_mb
        should_compress = not should_skip_compress
        if should_skip_compress and "skip-compress" not in markers:
            markers.append("skip-compress")
        if markers:
            row["markers"] = markers

        if category == "office":
            model_id = "prebuilt-invoice"
            file_type = "invoice"
        elif page_count == 1 and page_ratio is not None and page_ratio > _RECEIPT_RATIO_THRESHOLD:
            model_id = "prebuilt-receipt"
            file_type = "receipt"
        else:
            model_id = "prebuilt-invoice"
            file_type = "invoice"
        if should_compress:
            compress_task = asyncio.create_task(
                asyncio.to_thread(
                    _compress_pdf_for_archive,
                    source_path,
                    dest_dir=archive_root / category,
                    dpi=300,
                    name_suffix=batch.batch_id[:8],
                )
            )
        else:
            compress_task = asyncio.create_task(
                asyncio.to_thread(
                    _copy_pdf_for_archive_without_compression,
                    source_path,
                    dest_dir=archive_root / category,
                    name_suffix=batch.batch_id[:8],
                )
            )
        extract_task: asyncio.Task[Any] | None = None
        if should_extract:
            if use_compressed_for_di:
                async def _extract_after_compress() -> Any:
                    """Wait for compressed artifact and then run DI on compressed path."""

                    compressed_input = await compress_task
                    return await self._run_extract_with_controls(
                        pdf_path=Path(compressed_input),
                        model_id=model_id,
                        return_fields=(category == "office"),
                        extract_semaphore=extract_semaphore,
                        extract_rate_limiter=extract_rate_limiter,
                    )

                extract_task = asyncio.create_task(_extract_after_compress())
            else:
                extract_task = asyncio.create_task(
                    self._run_extract_with_controls(
                        pdf_path=source_path,
                        model_id=model_id,
                        return_fields=(category == "office"),
                        extract_semaphore=extract_semaphore,
                        extract_rate_limiter=extract_rate_limiter,
                    )
                )

        try:
            compressed_path = await compress_task
            row["preview_path"] = str(compressed_path)
        except Exception as exc:
            row["archive_error"] = str(exc)

        if extract_task is not None:
            try:
                extract_output = await extract_task
                if category == "office":
                    azure_result, office_fields = extract_output
                    self._fill_office_row(
                        row,
                        azure_result,
                        office_fields,
                        batch_metadata=batch.metadata,
                        batch_out_dir=archive_root.parent,
                    )
                else:
                    self._fill_daily_row(row, extract_output, file_type=file_type)
            except Exception as exc:
                row["extract_error"] = str(exc)

        preview_path_raw = row.get("preview_path")
        if isinstance(preview_path_raw, str) and preview_path_raw.strip():
            try:
                organized_path = _copy_pdf_to_organized_dir(
                    compressed_pdf=Path(preview_path_raw),
                    organized_root=organized_root,
                    category=category,
                    run_date=batch.run_date or "",
                    extracted_result=row.get("result") or {},
                )
                row["organized_path"] = str(organized_path)
            except Exception as exc:
                row["organize_error"] = str(exc)

        return row

    def _process_one_file(self, **_: Any) -> dict[str, Any]:
        """Legacy sync hook retained for timeout tests that monkeypatch this method."""

        raise RuntimeError("_process_one_file is only a legacy monkeypatch hook; use _process_one_file_async")

    def _fill_daily_row(self, row: dict[str, Any], azure_result: dict[str, Any], *, file_type: str) -> None:
        """Map Azure DI fields into daily review row contract; file_type drives tax_id extraction."""

        store_name = azure_result.get("store_name")
        if isinstance(store_name, str) and store_name.strip():
            row["result"]["store_name"] = store_name.splitlines()[0].strip()
        row["result"]["brutto"] = azure_result.get("brutto")
        row["result"]["netto"] = azure_result.get("netto")
        row["result"]["total_tax"] = azure_result.get("total_tax")
        row["result"]["file_type"] = file_type
        row["score"]["store_name"] = azure_result.get("confidence_store_name")
        row["score"]["brutto"] = azure_result.get("confidence_brutto")
        row["score"]["netto"] = azure_result.get("confidence_netto")
        row["score"]["total_tax"] = azure_result.get("confidence_total_tax")
        if file_type == "receipt":
            row["result"]["tax_id"] = "Beleg"
        else:
            row["result"]["bill_id"] = azure_result.get("bill_id")
            row["score"]["bill_id"] = azure_result.get("confidence_bill_id")

    def _fill_office_row(
        self,
        row: dict[str, Any],
        azure_result: dict[str, Any],
        office_fields: dict[str, Any],
        *,
        batch_metadata: dict[str, Any] | None,
        batch_out_dir: Path,
    ) -> None:
        """Map Azure invoice fields and optional Office semantics into review row."""

        row["result"]["brutto"] = azure_result.get("brutto")
        row["result"]["netto"] = azure_result.get("netto")
        row["result"]["tax_id"] = azure_result.get("invoice_id")
        row["score"]["brutto"] = azure_result.get("confidence_brutto")
        row["score"]["netto"] = azure_result.get("confidence_netto")
        row["score"]["tax_id"] = azure_result.get("confidence_invoice_id")

        try:
            distilled = _clean_invoice_fields(office_fields)
            self._persist_office_di_fields(
                batch_out_dir=batch_out_dir,
                row_id=str(row.get("row_id") or ""),
                distilled_fields=distilled,
            )
            office_info = _extract_office_semantics(distilled)
            row["result"]["type"] = office_info.get("purpose")
            row["result"]["sender"] = office_info.get("sender")
            row["result"]["receiver_name"] = office_info.get("receiver")
            row["result"]["receiver_ok"] = _resolve_receiver_ok(
                office_info,
                batch_metadata=batch_metadata,
            )
            row["result"]["receiver_address"] = office_info.get("receiver_address")
            row["result"]["receiver_address_ok"] = _resolve_receiver_address_ok(
                office_info,
                batch_metadata=batch_metadata,
            )
        except Exception as exc:
            row["semantic_error"] = str(exc)

    def _persist_office_di_fields(
        self,
        *,
        batch_out_dir: Path,
        row_id: str,
        distilled_fields: dict[str, Any],
    ) -> None:
        """Persist per-row distilled DI fields for prompt-tuning dataset collection."""

        normalized_row_id = row_id.strip() or "row-unknown"
        di_dir = batch_out_dir / "di_fields"
        di_dir.mkdir(parents=True, exist_ok=True)
        target_path = di_dir / f"{normalized_row_id}.json"
        target_path.write_text(
            json.dumps(distilled_fields, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )

    def _row_has_external_failure(self, row: dict[str, Any]) -> bool:
        """Return whether row contains extraction/semantic external-call failures."""

        return bool(row.get("error") or row.get("extract_error") or row.get("semantic_error"))

    def _row_status(self, row: dict[str, Any]) -> str:
        """Map one backend row payload into the persisted input status contract."""

        if row.get("skip_reason"):
            return "skipped"
        return "failed" if self._row_has_external_failure(row) else "extracted"

    def _row_error(self, row: dict[str, Any]) -> str | None:
        """Extract the canonical per-file error message from processing row payload."""

        error_value = row.get("error") or row.get("extract_error") or row.get("semantic_error")
        if error_value is None:
            return None
        text = str(error_value).strip()
        return text or None

    def _build_file_done_event(self, row: dict[str, Any]) -> dict[str, Any]:
        """Build worker callback payload for one completed file processing row."""

        return {
            "row_id": str(row.get("row_id") or ""),
            "filename": str(row.get("filename") or ""),
            "category": str(row.get("category") or ""),
            "status": self._row_status(row),
            "error": self._row_error(row),
        }

    async def merge_batch(self, batch: BatchRecord, payload: dict[str, Any]) -> dict[str, Any]:
        """Run real local Excel merge and return merged file artifact metadata."""

        out_dir = self.root / batch.batch_id
        out_dir.mkdir(parents=True, exist_ok=True)

        monthly_excel = self._resolve_monthly_excel_path(
            batch=batch,
            payload=payload,
            out_dir=out_dir,
        )
        append_mode = str(payload.get("mode", "overwrite")) == "append"

        validated_excel = out_dir / f"validated_for_merge_{int(datetime.now().timestamp())}.xlsx"
        if batch.batch_type.value == "daily":
            self._write_daily_validated_excel(batch, validated_excel)
            merged_excel = merge_daily(
                validated_excel,
                monthly_excel,
                out_dir=out_dir,
                append=append_mode,
            )
        else:
            self._write_office_validated_excel(batch, validated_excel)
            merged_excel = merge_office(
                validated_excel,
                monthly_excel,
                out_dir=out_dir,
                append=append_mode,
            )

        merge_summary_path = out_dir / "merge_summary.json"
        merge_summary_path.write_text(
            json.dumps(
                {
                    "batch_id": batch.batch_id,
                    "mode": payload.get("mode", "overwrite"),
                    "monthly_excel_path": str(monthly_excel.resolve()),
                    "validated_excel_path": str(validated_excel.resolve()),
                    "merged_excel_path": str(merged_excel.resolve()),
                    "merged_excel_download_path": f"/v1/batches/{batch.batch_id}/merge-output/download",
                    "review_rows_count": len(batch.review_rows),
                    "generated_at": datetime.now(UTC).isoformat(),
                },
                ensure_ascii=False,
                indent=2,
            ),
            encoding="utf-8",
        )
        merged_abs_path = str(merged_excel.resolve())
        merged_download_path = f"/v1/batches/{batch.batch_id}/merge-output/download"
        return {
            "merge_summary_path": str(merge_summary_path.resolve()),
            "validated_excel_path": str(validated_excel.resolve()),
            "merged_excel_path": merged_download_path,
            "output_path": merged_download_path,
            "merged_excel_abs_path": merged_abs_path,
            "output_abs_path": merged_abs_path,
            "merge_mode": str(payload.get("mode", "overwrite")),
        }

    def _resolve_monthly_excel_path(
        self,
        *,
        batch: BatchRecord,
        payload: dict[str, Any],
        out_dir: Path,
    ) -> Path:
        """Resolve merge target monthly workbook path, defaulting to batch-local auto template path."""

        raw_monthly_excel_path = payload.get("monthly_excel_path")
        if isinstance(raw_monthly_excel_path, str) and raw_monthly_excel_path.strip():
            return Path(raw_monthly_excel_path.strip())
        return out_dir / "merge_source" / f"auto_{batch.batch_type.value}_monthly.xlsx"

    def _write_daily_validated_excel(self, batch: BatchRecord, out_path: Path) -> None:
        """Build daily validated workbook from review rows using legacy mapper style logic."""

        items = []
        for row in batch.review_rows:
            category = str(row.get("category") or "").lower()
            if category not in {"bar", "zbon"}:
                continue
            result = dict(row.get("result") or {})
            if not result.get("run_date") and batch.run_date:
                result["run_date"] = batch.run_date
            items.append(
                {
                    "category": category,
                    "filename": row.get("filename"),
                    "result": result,
                    "score": row.get("score") or {},
                    "preview_path": row.get("preview_path") or row.get("preview_url"),
                }
            )

        if not items:
            raise ValueError("No daily review rows available for merge")
        temp_json_path = out_path.with_suffix(".rows.json")
        temp_json_path.write_text(json.dumps(items, ensure_ascii=False, indent=2), encoding="utf-8")
        map_daily_json_to_excel(
            temp_json_path,
            excel_path=out_path,
            config_path=Path("tests") / "config.json",
        )

    def _write_office_validated_excel(self, batch: BatchRecord, out_path: Path) -> None:
        """Build office validated workbook from current review rows."""

        headers = [
            "Datum",
            "Type",
            "Rechnung Name",
            "Brutto",
            "Netto",
            "Steuernummer",
            "Is Receiver OK",
            "need review",
            "Rechnung Scannen",
        ]
        wb = Workbook()
        ws = wb.active
        ws.title = "Office"
        ws.append(headers)

        row_idx = 2
        for row in batch.review_rows:
            if str(row.get("category") or "").lower() != "office":
                continue
            result = row.get("result") or {}
            datum = normalize_date(result.get("run_date") or batch.run_date) or (batch.run_date or "")
            excel_row = [
                datum,
                result.get("type"),
                result.get("sender"),
                result.get("brutto"),
                result.get("netto"),
                result.get("tax_id"),
                result.get("receiver_ok"),
                bool(row.get("need review", False)),
                row.get("preview_path") or row.get("preview_url"),
            ]
            ws.append(excel_row)
            write_datum_cell(ws.cell(row=row_idx, column=1), datum)
            scan_col_idx = headers.index("Rechnung Scannen")
            link = _to_excel_hyperlink(excel_row[scan_col_idx])
            if link:
                link_cell = ws.cell(row=row_idx, column=scan_col_idx + 1)
                link_cell.value = "check pdf"
                link_cell.hyperlink = link
            row_idx += 1

        if row_idx == 2:
            raise ValueError("No office review rows available for merge")

        wb.save(out_path)





