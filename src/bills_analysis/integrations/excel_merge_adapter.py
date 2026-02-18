from __future__ import annotations

import shutil
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from bills_analysis.excel_ops import (
    merge_validated_row,
    normalize_date,
    normalize_header,
    parse_datum,
    write_datum_cell,
)


def _cell_has_value(value: Any) -> bool:
    """Check whether an Excel cell value should be treated as non-empty."""

    if value is None:
        return False
    if isinstance(value, str):
        return value.strip() != ""
    return True


def _load_single_row(path: Path) -> tuple[list[str], list[Any]]:
    """Load first data row and headers from validated workbook."""

    wb = load_workbook(path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    if ws.max_row < 2:
        raise ValueError("Validated Excel must contain exactly one data row.")
    row_values = [cell.value for cell in ws[2]]
    return [str(h).strip() if h is not None else "" for h in headers], row_values


def _find_row_by_datum(ws: Any, datum: str) -> int | None:
    """Find monthly workbook row index by normalized Datum value."""

    target = normalize_date(datum) or str(datum).strip()
    for row_idx in range(2, ws.max_row + 1):
        cell_value = ws.cell(row=row_idx, column=1).value
        cell_text = normalize_date(cell_value) or str(cell_value).strip()
        if cell_text == target:
            return row_idx
    return None


def _build_daily_template_headers(*, max_expense_rows: int = 5) -> list[str]:
    """Build canonical daily monthly-template headers for first-time workbook creation."""

    headers = ["Datum", "Umsatz Brutto", "Umsatz Netto", "Wie viel Rechnungen"]
    for idx in range(1, max_expense_rows + 1):
        headers.extend(
            [
                f"Ausgabe {idx} Name",
                f"Ausgabe {idx} Brutto",
                f"Ausgabe {idx} Netto",
            ]
        )
    return headers


def _ensure_daily_monthly_template(monthly_xlsx: Path) -> None:
    """Create a blank daily monthly workbook with canonical headers when target is missing."""

    if monthly_xlsx.exists():
        return
    monthly_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Daily"
    ws.append(_build_daily_template_headers())
    wb.save(monthly_xlsx)


def _build_office_template_headers() -> list[str]:
    """Build canonical office monthly-template headers for first-time workbook creation."""

    return [
        "Datum",
        "Type",
        "Rechnung Name",
        "Brutto",
        "Netto",
        "Steuernummer",
        "Is Receiver OK",
        "Is Receiver Address OK",
        "Rechnung Scannen",
    ]


def _ensure_office_monthly_template(monthly_xlsx: Path) -> None:
    """Create a blank office monthly workbook with canonical headers when target is missing."""

    if monthly_xlsx.exists():
        return
    monthly_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Office"
    ws.append(_build_office_template_headers())
    wb.save(monthly_xlsx)


def _sort_daily_rows_by_datum(ws: Any) -> None:
    """Sort data rows by Datum ascending; unparseable Datum values are placed last."""

    rows: list[list[Any]] = []
    for row_idx in range(2, ws.max_row + 1):
        row_values = [ws.cell(row=row_idx, column=col_idx).value for col_idx in range(1, ws.max_column + 1)]
        if any(_cell_has_value(value) for value in row_values):
            rows.append(row_values)

    def _sort_key(row_values: list[Any]) -> tuple[int, date, str]:
        datum_raw = row_values[0] if row_values else None
        datum_norm = normalize_date(datum_raw) or str(datum_raw or "").strip()
        datum_parsed = parse_datum(datum_norm)
        if datum_parsed is not None:
            return (0, datum_parsed, "")
        return (1, date.max, datum_norm.lower())

    rows.sort(key=_sort_key)
    if ws.max_row >= 2:
        ws.delete_rows(2, ws.max_row - 1)
    for row_values in rows:
        ws.append(row_values)
        current_row = ws.max_row
        if row_values:
            write_datum_cell(ws.cell(row=current_row, column=1), row_values[0])


def _build_output_workbook_path(
    *,
    out_dir: Path,
    validated_xlsx: Path,
    monthly_xlsx: Path,
) -> Path:
    """Build a unique merge output path and guard against input/output path collisions."""

    out_path = out_dir / f"full_result_{datetime.now().strftime('%Y%m%d_%H%M%S_%f')}.xlsx"
    if out_path.resolve() == validated_xlsx.resolve():
        raise ValueError("输出路径与已校验文件相同，请指定不同的输出目录。")
    if out_path.resolve() == monthly_xlsx.resolve():
        raise ValueError("输出路径与全量文件相同，请指定不同的输出目录。")
    return out_path


def merge_daily_excel(
    validated_xlsx: Path,
    monthly_xlsx: Path,
    *,
    out_dir: Path | None = None,
    append: bool = False,
) -> Path:
    """Merge daily validated one-row Excel into monthly workbook with overwrite/append mode."""

    validated_headers, validated_row = _load_single_row(validated_xlsx)
    if not validated_headers or validated_headers[0] != "Datum":
        raise ValueError("Validated Excel must have 'Datum' as the first column.")
    datum = str(validated_row[0]).strip() if validated_row else ""
    if not datum:
        raise ValueError("Validated Excel has empty Datum.")

    out_dir = out_dir or monthly_xlsx.parent
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = _build_output_workbook_path(
        out_dir=out_dir,
        validated_xlsx=validated_xlsx,
        monthly_xlsx=monthly_xlsx,
    )

    _ensure_daily_monthly_template(monthly_xlsx)
    shutil.copy2(monthly_xlsx, out_path)
    wb = load_workbook(out_path)
    ws = wb.active
    monthly_headers = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[1]]
    header_to_col = {name: idx + 1 for idx, name in enumerate(monthly_headers)}

    updates, _missing_headers = merge_validated_row(validated_headers, validated_row, monthly_headers)
    if append:
        target_row = ws.max_row + 1
    else:
        target_row = _find_row_by_datum(ws, datum)
        if target_row is None:
            target_row = ws.max_row + 1

    for header, value in updates.items():
        col_idx = header_to_col.get(header)
        if col_idx is None:
            continue
        cell = ws.cell(row=target_row, column=col_idx)
        if normalize_header(header) == normalize_header("Datum"):
            write_datum_cell(cell, value)
        else:
            cell.value = value

    if "Datum" in header_to_col:
        write_datum_cell(ws.cell(row=target_row, column=header_to_col["Datum"]), datum)
    _sort_daily_rows_by_datum(ws)
    wb.save(out_path)
    return out_path


def _load_all_rows(path: Path) -> tuple[list[str], list[list[Any]], list[list[str | None]]]:
    """Load all data rows and hyperlink targets from validated workbook."""

    wb = load_workbook(path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    rows: list[list[Any]] = []
    links: list[list[str | None]] = []
    for row_idx in range(2, ws.max_row + 1):
        row_cells = ws[row_idx]
        rows.append([cell.value for cell in row_cells])
        links.append([cell.hyperlink.target if cell.hyperlink is not None else None for cell in row_cells])
    return [str(h).strip() if h is not None else "" for h in headers], rows, links


def merge_office_excel(
    validated_xlsx: Path,
    monthly_xlsx: Path,
    *,
    out_dir: Path | None = None,
    append: bool = False,
) -> Path:
    """Merge office validated Excel into monthly workbook in overwrite/append mode."""

    validated_headers, validated_rows, validated_links = _load_all_rows(validated_xlsx)
    if not validated_headers or normalize_header(validated_headers[0]) != normalize_header("Datum"):
        raise ValueError("Validated Excel must have 'Datum' as the first column.")
    if not validated_rows:
        raise ValueError("Validated Excel has no data rows.")

    out_dir = out_dir or monthly_xlsx.parent
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = _build_output_workbook_path(
        out_dir=out_dir,
        validated_xlsx=validated_xlsx,
        monthly_xlsx=monthly_xlsx,
    )

    _ensure_office_monthly_template(monthly_xlsx)
    shutil.copy2(monthly_xlsx, out_path)
    wb = load_workbook(out_path)
    ws = wb.active
    existing_monthly_last_row = ws.max_row
    consumed_existing_rows: set[int] = set()

    def _find_row_by_datum(datum: Any) -> int | None:
        """Resolve one existing target row by Datum for overwrite mode without reusing appended rows."""

        target = normalize_date(datum) or str(datum).strip()
        for row_idx in range(2, existing_monthly_last_row + 1):
            if row_idx in consumed_existing_rows:
                continue
            cell_value = ws.cell(row=row_idx, column=1).value
            cell_norm = normalize_date(cell_value) or str(cell_value).strip()
            if cell_norm == target:
                consumed_existing_rows.add(row_idx)
                return row_idx
        return None

    for row, row_links in zip(validated_rows, validated_links):
        if any(normalize_header(h) == normalize_header("need review") for h in validated_headers):
            filtered_headers = []
            filtered_row = []
            filtered_links: list[str | None] = []
            for h, v in zip(validated_headers, row):
                if normalize_header(h) == normalize_header("need review"):
                    continue
                filtered_headers.append(h)
                filtered_row.append(v)
            for h, link in zip(validated_headers, row_links):
                if normalize_header(h) == normalize_header("need review"):
                    continue
                filtered_links.append(link)
        else:
            filtered_headers = validated_headers
            filtered_row = row
            filtered_links = row_links

        if append:
            ws.append(filtered_row)
            row_idx = ws.max_row
            if filtered_row:
                write_datum_cell(ws.cell(row=row_idx, column=1), filtered_row[0])
            for col_idx, link in enumerate(filtered_links, start=1):
                if link:
                    ws.cell(row=row_idx, column=col_idx).hyperlink = link
            continue

        datum_val = filtered_row[0] if filtered_row else ""
        target_row = _find_row_by_datum(datum_val)
        if target_row is None:
            ws.append(filtered_row)
            continue

        for col_idx, (value, link) in enumerate(zip(filtered_row, filtered_links), start=1):
            if col_idx == 1:
                write_datum_cell(ws.cell(row=target_row, column=col_idx), value)
                continue
            cell = ws.cell(row=target_row, column=col_idx, value=value)
            cell.hyperlink = link

    wb.save(out_path)
    return out_path
