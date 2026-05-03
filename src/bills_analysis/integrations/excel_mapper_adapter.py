from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import PatternFill

from bills_analysis.excel_ops import (
    build_rows_with_meta,
    low_confidence_fields,
    normalize_date,
    threshold_for,
    to_score,
    write_datum_cell,
)
from bills_analysis.integrations.app_config import resolve_app_config_path


def load_results(path: Path) -> list[dict[str, Any]]:
    """Load pipeline results from JSON array or line-delimited JSON file."""

    raw = path.read_text(encoding="utf-8").strip()
    if not raw:
        return []
    if raw.lstrip().startswith("["):
        data = json.loads(raw)
        if isinstance(data, list):
            return data
        raise ValueError("Results JSON is not a list.")
    return [json.loads(line) for line in raw.splitlines() if line.strip()]


def load_json_object(path: Path, *, empty_message: str) -> dict[str, Any]:
    """Load a JSON object from file and validate the top-level type."""

    raw = path.read_text(encoding="utf-8").strip()
    if not raw:
        raise ValueError(empty_message)
    data = json.loads(raw)
    if not isinstance(data, dict):
        raise ValueError("JSON must be an object.")
    return data


def to_link(value: Any, base_dir: Path) -> str | None:
    """Convert local/remote preview path values into Excel hyperlink targets."""

    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    if text.startswith("http://") or text.startswith("https://"):
        return text
    path = Path(text)
    if not path.is_absolute():
        path = (base_dir / path).resolve()
    try:
        return path.as_uri()
    except ValueError:
        return str(path)


def map_daily_json_to_excel(
    json_path: Path,
    *,
    excel_path: Path | None = None,
    config_path: Path | None = None,
) -> Path:
    """Map daily results JSON to one-row Excel with low-confidence highlights."""

    out_path = excel_path or json_path.with_suffix(".xlsx")
    items = load_results(json_path)
    thresholds_path = config_path or resolve_app_config_path()
    thresholds = load_json_object(thresholds_path, empty_message=f"Empty thresholds file: {thresholds_path}")
    rows, zbon_files_by_date = build_rows_with_meta(items, thresholds)
    if not rows:
        raise ValueError("No rows generated.")
    first = rows[0]

    def _max_pages(default: int = 4) -> int:
        """Resolve max pages threshold from config."""

        value = thresholds.get("max_pages", default)
        try:
            return int(value)
        except (TypeError, ValueError):
            return default

    datum = first.get("Datum") or "UNKNOWN"
    need_review = bool(first.get("need review"))
    for item in items:
        result = item.get("result") or {}
        run_date = normalize_date(result.get("run_date")) or "UNKNOWN"
        if run_date != datum:
            continue
        page_count = item.get("page_count")
        if isinstance(page_count, int) and page_count > _max_pages():
            need_review = True
            break
    first["need review"] = need_review
    headers = list(first.keys())

    orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"
    ws.append(headers)
    ws.append([first.get(h) for h in headers])

    header_to_col = {name: idx + 1 for idx, name in enumerate(headers)}
    data_row_idx = 2
    low_headers = set()

    for item in items:
        result = item.get("result") or {}
        run_date = normalize_date(result.get("run_date")) or "UNKNOWN"
        if run_date != datum:
            continue
        category = str(item.get("category") or "").strip().lower()
        if category != "zbon":
            continue
        low_fields = low_confidence_fields(result, item.get("score") or {}, thresholds)
        if "brutto" in low_fields:
            low_headers.add("Umsatz Brutto")
        if "netto" in low_fields:
            low_headers.add("Umsatz Netto")

    zbon_files = zbon_files_by_date.get(datum, [])
    for idx, fname in enumerate(zbon_files, start=1):
        for item in items:
            if str(item.get("filename") or "") != fname:
                continue
            low_fields = low_confidence_fields(item.get("result") or {}, item.get("score") or {}, thresholds)
            if "store_name" in low_fields:
                low_headers.add(f"Ausgabe {idx} Name")
            if "brutto" in low_fields:
                low_headers.add(f"Ausgabe {idx} Brutto")
            if "netto" in low_fields:
                low_headers.add(f"Ausgabe {idx} Netto")
            break

    for header in low_headers:
        col = header_to_col.get(header)
        if col is not None:
            ws.cell(row=data_row_idx, column=col).fill = orange_fill
    if need_review:
        col = header_to_col.get("need review")
        if col is not None:
            ws.cell(row=data_row_idx, column=col).fill = orange_fill

    preview_map = {}
    for item in items:
        filename = str(item.get("filename") or "")
        preview_path = item.get("preview_path")
        if filename:
            preview_map[filename] = preview_path
    link_row_idx = data_row_idx + 1
    zbon_files = zbon_files_by_date.get(datum, [])
    for idx, fname in enumerate(zbon_files, start=1):
        if idx > 5:
            break
        preview = preview_map.get(fname)
        link = to_link(preview, json_path.parent) if preview else None
        if not link:
            continue
        col = header_to_col.get(f"Ausgabe {idx} Name")
        if col is None:
            continue
        cell = ws.cell(row=link_row_idx, column=col)
        cell.value = "check pdf"
        cell.hyperlink = link

    wb.save(out_path)
    return out_path


def map_office_json_to_excel(
    json_path: Path,
    *,
    excel_path: Path | None = None,
    config_path: Path | None = None,
) -> Path:
    """Map office results JSON to multi-row Excel with reliability highlights."""

    out_path = excel_path or json_path.with_suffix(".xlsx")
    items = load_results(json_path)
    config_source = config_path or resolve_app_config_path()
    config = load_json_object(config_source, empty_message=f"Empty config file: {config_source}")

    headers = [
        "Datum",
        "Type",
        "Rechnung Name",
        "Brutto",
        "Netto",
        "Steuernummer",
        "Is Receiver OK",
        "need review",
        "Rechnung Scannen",
    ]
    rows: list[list[Any]] = []
    row_meta: list[dict[str, Any]] = []

    for item in items:
        category = str(item.get("category") or "").strip().lower()
        if category != "office":
            continue
        result = item.get("result") or {}
        rows.append(
            [
                normalize_date(result.get("run_date")) or result.get("run_date"),
                result.get("type"),
                result.get("sender"),
                result.get("brutto"),
                result.get("netto"),
                result.get("tax_id"),
                result.get("receiver_ok"),
                False,
                result.get("preview_path") or item.get("preview_path"),
            ]
        )
        row_meta.append(
            {
                "brutto": result.get("brutto"),
                "netto": result.get("netto"),
                "total_tax": result.get("total_tax"),
                "receiver_ok": result.get("receiver_ok"),
                "preview_path": result.get("preview_path") or item.get("preview_path"),
                "score": item.get("score") or {},
            }
        )

    if not rows:
        raise ValueError("No OFFICE rows found.")

    wb = Workbook()
    ws = wb.active
    ws.title = "Office"
    ws.append(headers)
    orange = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
    for row_idx, row in enumerate(rows, start=2):
        ws.append(row)
        write_datum_cell(ws.cell(row=row_idx, column=1), row[0])
        meta = row_meta[row_idx - 2]
        score = meta.get("score") or {}
        need_review = False
        for field, header in [("brutto", "Brutto"), ("netto", "Netto")]:
            score_val = to_score(score.get(field))
            if score_val is None or score_val < threshold_for(field, config):
                col = headers.index(header) + 1
                ws.cell(row=row_idx, column=col).fill = orange
                need_review = True
        tax_val = row[headers.index("Steuernummer")]
        if tax_val in (None, "", "None"):
            col = headers.index("Steuernummer") + 1
            ws.cell(row=row_idx, column=col).fill = orange
            need_review = True
        if meta.get("receiver_ok") is not True:
            col = headers.index("Is Receiver OK") + 1
            ws.cell(row=row_idx, column=col).fill = orange
            need_review = True
        col = headers.index("need review") + 1
        ws.cell(row=row_idx, column=col, value=need_review)
        if need_review:
            ws.cell(row=row_idx, column=col).fill = orange
        link = to_link(meta.get("preview_path"), json_path.parent)
        if link:
            col = headers.index("Rechnung Scannen") + 1
            cell = ws.cell(row=row_idx, column=col)
            cell.value = "check pdf"
            cell.hyperlink = link
    wb.save(out_path)
    return out_path
