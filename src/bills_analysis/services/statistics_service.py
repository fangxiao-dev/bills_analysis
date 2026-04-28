from __future__ import annotations
"""Monthly statistics aggregation from Daily/Bar and Office Excel workbooks."""

import re
from collections import defaultdict
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from bills_analysis.models.api_responses import (
    DailyExpenseRow,
    DailyStatisticsPoint,
    ExpenseBreakdownItem,
    MonthlyStatisticsResponse,
    OfficeStatisticsRow,
    OfficeTypeBreakdown,
    StatisticsSummary,
)

_AUSGABE_BRUTTO_RE = re.compile(r"^Ausgabe \d+ Brutto$", re.IGNORECASE)
_AUSGABE_NETTO_RE = re.compile(r"^Ausgabe \d+ Netto$", re.IGNORECASE)
UNCATEGORIZED = "Uncategorized"


class DuplicateManualExpenseTypesError(ValueError):
    """Raised when manual expense types duplicate Office workbook types."""

    def __init__(self, duplicate_types: list[str]) -> None:
        super().__init__("Manual expense types already exist in Office workbook")
        self.duplicate_types = duplicate_types


def _headers(ws: Worksheet) -> dict[str, int]:
    """Return a column-name to zero-based index map from the first row."""

    row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if row is None:
        return {}
    return {str(cell).strip(): idx for idx, cell in enumerate(row) if cell is not None}


def _require(headers: dict[str, int], names: list[str], label: str) -> None:
    """Raise ValueError listing any required column names not found in headers."""

    missing = [name for name in names if name not in headers]
    if missing:
        raise ValueError(f"{label} workbook is missing required columns: {', '.join(missing)}")


def _to_decimal(value: Any, warnings: list[str], label: str) -> Decimal:
    """Convert a cell value to Decimal, recording a warning and returning 0 on failure."""

    if value is None or value == "":
        return Decimal("0")
    text = str(value).strip()
    if text.endswith("€"):
        text = text[:-1].strip()
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        return Decimal(text)
    except InvalidOperation:
        warnings.append(f"Non-numeric value in {label}: {value!r}; treated as 0")
        return Decimal("0")


def _to_manual_decimal(value: Any, label: str) -> Decimal:
    """Convert manual input amount where comma or dot means decimal separator."""

    if value is None or value == "":
        raise ValueError(f"{label} is required")
    text = str(value).strip()
    if text.endswith("€"):
        text = text[:-1].strip()
    text = text.replace(" ", "")
    if "," in text and "." in text:
        raise ValueError(f"{label} must not contain thousands separators")
    if text.count(",") > 1 or text.count(".") > 1:
        raise ValueError(f"{label} must contain at most one decimal separator")
    text = text.replace(",", ".")
    try:
        return Decimal(text)
    except InvalidOperation as exc:
        raise ValueError(f"{label} must be numeric") from exc


def _date_text(value: Any) -> str | None:
    """Convert a cell date value to an ISO date string, or return None."""

    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    return text if text else None


def _money(value: Decimal) -> float:
    """Round to 2 decimal places and return as float for JSON output."""

    return float(value.quantize(Decimal("0.01")))


def _share(value: Decimal, total: Decimal) -> float:
    """Return a four-decimal share for type breakdowns."""

    if total == 0:
        return 0.0
    return float((value / total).quantize(Decimal("0.0001")))


def build_monthly_statistics(
    daily_xlsx: Path,
    office_xlsx: Path,
    *,
    manual_expense_rows: list[dict[str, Any]] | None = None,
    allow_duplicate_manual_types: bool = False,
) -> MonthlyStatisticsResponse:
    """Parse Daily/Bar and Office workbooks and return aggregated monthly statistics."""

    warnings: list[str] = []

    try:
        daily_wb = load_workbook(daily_xlsx, data_only=True)
    except Exception as exc:
        raise ValueError(f"Cannot read Daily workbook: {exc}") from exc

    try:
        office_wb = load_workbook(office_xlsx, data_only=True)
    except Exception as exc:
        raise ValueError(f"Cannot read Office workbook: {exc}") from exc

    daily_series, daily_expense_rows, revenue_total, daily_expense_total = _parse_daily(daily_wb.active, warnings)
    office_rows, office_by_type, office_total = _parse_office(
        office_wb.active,
        warnings,
        manual_expense_rows=manual_expense_rows or [],
        allow_duplicate_manual_types=allow_duplicate_manual_types,
    )
    profit = revenue_total - daily_expense_total - office_total
    expense_breakdown = _build_expense_breakdown(daily_expense_total, daily_expense_rows, office_by_type, office_total)

    return MonthlyStatisticsResponse(
        summary=StatisticsSummary(
            revenue_brutto=_money(revenue_total),
            daily_expense_brutto=_money(daily_expense_total),
            office_expense_brutto=_money(office_total),
            profit_brutto=_money(profit),
        ),
        daily_series=daily_series,
        office_by_type=office_by_type,
        office_rows=office_rows,
        expense_breakdown=expense_breakdown,
        daily_expense_rows=daily_expense_rows,
        warnings=warnings,
    )


def _parse_daily(ws: Worksheet, warnings: list[str]) -> tuple[list[DailyStatisticsPoint], list[DailyExpenseRow], Decimal, Decimal]:
    """Parse Daily sheet and return series, daily expense rows, revenue total, and expense total."""

    headers = _headers(ws)
    _require(headers, ["Datum", "Umsatz Brutto"], "Daily")
    expense_cols = [name for name in headers if _AUSGABE_BRUTTO_RE.match(name)]
    expense_netto_cols = [name for name in headers if _AUSGABE_NETTO_RE.match(name)]

    series: list[DailyStatisticsPoint] = []
    expense_by_date: dict[str, Decimal] = defaultdict(lambda: Decimal("0"))
    expense_netto_by_date: dict[str, Decimal] = defaultdict(lambda: Decimal("0"))
    revenue_total = Decimal("0")
    expense_total = Decimal("0")

    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(cell is None for cell in row):
            continue
        date_val = _date_text(row[headers["Datum"]])
        revenue = _to_decimal(row[headers["Umsatz Brutto"]], warnings, "Daily Umsatz Brutto")
        expense = sum((_to_decimal(row[headers[col]], warnings, col) for col in expense_cols), Decimal("0"))
        expense_netto = sum((_to_decimal(row[headers[col]], warnings, col) for col in expense_netto_cols), Decimal("0"))
        revenue_total += revenue
        expense_total += expense
        if date_val and expense > 0:
            if not expense_netto_cols:
                warnings.append("Daily workbook has Bar Ausgabe Brutto values but no Ausgabe N Netto columns")
            elif expense_netto == 0:
                warnings.append(f"Daily Bar Ausgabe Netto is empty or zero for {date_val}")
            expense_by_date[date_val] += expense
            if expense_netto_cols:
                expense_netto_by_date[date_val] += expense_netto
        series.append(
            DailyStatisticsPoint(
                date=date_val or "",
                revenue_brutto=_money(revenue),
                daily_expense_brutto=_money(expense),
                profit_before_office_brutto=_money(revenue - expense),
            )
        )

    daily_expense_rows = [
        DailyExpenseRow(
            date=date_key,
            brutto=_money(total),
            netto=_money(expense_netto_by_date[date_key]) if expense_netto_cols else None,
        )
        for date_key, total in sorted(expense_by_date.items(), key=lambda item: item[0])
        if total > 0
    ]

    return series, daily_expense_rows, revenue_total, expense_total


def _parse_office(
    ws: Worksheet,
    warnings: list[str],
    *,
    manual_expense_rows: list[dict[str, Any]],
    allow_duplicate_manual_types: bool,
) -> tuple[list[OfficeStatisticsRow], list[OfficeTypeBreakdown], Decimal]:
    """Parse the active Office sheet and return rows, type breakdowns, and total."""

    headers = _headers(ws)
    _require(headers, ["Type", "Brutto"], "Office")

    has_datum = "Datum" in headers
    has_name = "Rechnung Name" in headers
    has_netto = "Netto" in headers
    rows: list[OfficeStatisticsRow] = []
    type_totals: dict[str, Decimal] = defaultdict(lambda: Decimal("0"))
    type_counts: dict[str, int] = defaultdict(int)
    office_total = Decimal("0")

    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(cell is None for cell in row):
            continue

        raw_type = row[headers["Type"]]
        office_type = str(raw_type).strip() if raw_type is not None else ""
        if not office_type:
            office_type = UNCATEGORIZED

        brutto = _to_decimal(row[headers["Brutto"]], warnings, "Office Brutto")
        netto = _to_decimal(row[headers["Netto"]], warnings, "Office Netto") if has_netto else None
        date_val = _date_text(row[headers["Datum"]]) if has_datum else None
        name_val = None
        if has_name:
            raw_name = row[headers["Rechnung Name"]]
            name_val = str(raw_name).strip() if raw_name is not None else None
            if not name_val:
                name_val = None

        rows.append(
            OfficeStatisticsRow(
                date=date_val,
                type=office_type,
                name=name_val,
                brutto=_money(brutto),
                netto=_money(netto) if netto is not None else None,
            )
        )
        type_totals[office_type] += brutto
        type_counts[office_type] += 1
        office_total += brutto

    _append_manual_expense_rows(
        manual_expense_rows,
        rows=rows,
        type_totals=type_totals,
        type_counts=type_counts,
        allow_duplicate_manual_types=allow_duplicate_manual_types,
    )
    office_total = sum(type_totals.values(), Decimal("0"))

    by_type = sorted(
        [
            OfficeTypeBreakdown(
                type=office_type,
                brutto=_money(total),
                count=type_counts[office_type],
                share=_share(total, office_total),
            )
            for office_type, total in type_totals.items()
        ],
        key=lambda breakdown: breakdown.brutto,
        reverse=True,
    )

    return rows, by_type, office_total


def _append_manual_expense_rows(
    manual_rows: list[dict[str, Any]],
    *,
    rows: list[OfficeStatisticsRow],
    type_totals: dict[str, Decimal],
    type_counts: dict[str, int],
    allow_duplicate_manual_types: bool,
) -> None:
    """Append confirmed manual Ausgabe rows into Office aggregation state."""

    if not manual_rows:
        return

    canonical_by_key = {office_type.strip().casefold(): office_type for office_type in type_totals}
    duplicates: list[str] = []
    for item in manual_rows:
        manual_type = str(item.get("type") or "").strip()
        if manual_type and manual_type.casefold() in canonical_by_key:
            duplicates.append(manual_type)
    if duplicates and not allow_duplicate_manual_types:
        deduped = list(dict.fromkeys(duplicates))
        raise DuplicateManualExpenseTypesError(deduped)

    for item in manual_rows:
        manual_type = str(item.get("type") or "").strip()
        if not manual_type:
            raise ValueError("manual expense type is required")
        office_type = canonical_by_key.get(manual_type.casefold(), manual_type)
        brutto = _to_manual_decimal(item.get("brutto"), f"Manual {manual_type} Brutto")
        netto_value = item.get("netto")
        netto = _to_manual_decimal(netto_value, f"Manual {manual_type} Netto") if netto_value not in (None, "") else None
        rows.append(
            OfficeStatisticsRow(
                date=None,
                type=office_type,
                name=office_type,
                brutto=_money(brutto),
                netto=_money(netto) if netto is not None else None,
            )
        )
        type_totals[office_type] += brutto
        type_counts[office_type] += 1
        canonical_by_key.setdefault(office_type.casefold(), office_type)


def _build_expense_breakdown(
    daily_expense_total: Decimal,
    daily_expense_rows: list[DailyExpenseRow],
    office_by_type: list[OfficeTypeBreakdown],
    office_total: Decimal,
) -> list[ExpenseBreakdownItem]:
    """Combine Bar Ausgabe and Office type totals for the expense breakdown chart."""

    total_expense = daily_expense_total + office_total
    items: list[ExpenseBreakdownItem] = []

    if daily_expense_total != 0 or daily_expense_rows:
        items.append(
            ExpenseBreakdownItem(
                category="Bar Ausgabe",
                source="daily_bar",
                brutto=_money(daily_expense_total),
                count=len(daily_expense_rows),
                share=_share(daily_expense_total, total_expense),
            )
        )

    items.extend(
        ExpenseBreakdownItem(
            category=item.type,
            source="office",
            brutto=item.brutto,
            count=item.count,
            share=_share(Decimal(str(item.brutto)), total_expense),
        )
        for item in office_by_type
    )

    return sorted(items, key=lambda item: item.brutto, reverse=True)
